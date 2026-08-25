import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_financial_model():
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Financial Analysis (Main Sheet)
    # ----------------------------------------------------
    ws = wb.active
    ws.title = "Financial Analysis"
    ws.views.sheetView[0].showGridLines = True

    # ----------------------------------------------------
    # Sheet 2: Depreciation
    # ----------------------------------------------------
    ws_dep = wb.create_sheet("Depreciation")
    ws_dep.views.sheetView[0].showGridLines = True

    # Colors
    color_title_bg = "F2EBEF"  # Soft pink/gray
    color_main_sheet_bg = "5E2750"  # Dark purple
    color_section_bg = "F9F9F9"  # Very light gray for section labels

    # Fonts
    font_title = Font(name="Calibri", size=16, bold=True)
    font_main_sheet = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_header_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_italic = Font(name="Calibri", size=11, italic=True)
    font_bold_italic = Font(name="Calibri", size=11, bold=True, italic=True)

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_vertical_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    dark_thin_side = Side(border_style="thin", color="000000")
    double_bottom_side = Side(border_style="double", color="000000")

    border_all_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_bottom_double = Border(bottom=double_bottom_side, left=thin_side, right=thin_side, top=thin_side)
    border_dark_bottom = Border(bottom=dark_thin_side, top=dark_thin_side)

    # Number Formats
    fmt_decimal = "#,##0.00;(#,##0.00);\"-\""
    fmt_percent = "0.00%"
    fmt_rate = "0%"

    # ----------------------------------------------------
    # POPULATE DEPRECIATION SHEET
    # ----------------------------------------------------
    # Title Row
    ws_dep.merge_cells("A1:AC1")
    ws_dep["A1"] = "Depreciation"
    ws_dep["A1"].font = font_title
    ws_dep["A1"].alignment = align_left
    ws_dep["A1"].fill = PatternFill(start_color=color_title_bg, end_color=color_title_bg, fill_type="solid")
    ws_dep.row_dimensions[1].height = 35

    # Main Sheet Header
    ws_dep.merge_cells("B2:C2")
    ws_dep["B2"] = "Main Sheet"
    ws_dep["B2"].font = font_main_sheet
    ws_dep["B2"].alignment = align_center
    ws_dep["B2"].fill = PatternFill(start_color=color_main_sheet_bg, end_color=color_main_sheet_bg, fill_type="solid")
    ws_dep.row_dimensions[2].height = 20

    ws_dep["A2"] = "(Figures in Rs. Crores)"
    ws_dep["A2"].font = font_italic

    # Headers: Year and FY
    ws_dep["A3"] = "Year"
    ws_dep["A3"].font = font_header_bold
    ws_dep["A3"].alignment = align_left

    ws_dep["A4"] = "FY"
    ws_dep["A4"].font = font_header_bold
    ws_dep["A4"].alignment = align_left

    years_num = list(range(1, 27))
    years_ending = [
        "2025-26", "2026-27", "2027-28", "2028-29", "2029-30", "2030-31",
        "2031-32", "2032-33", "2033-34", "2034-35", "2035-36", "2036-37",
        "2037-38", "2038-39", "2039-40", "2040-41", "2041-42", "2042-43",
        "2043-44", "2044-45", "2045-46", "2046-47", "2047-48", "2048-49",
        "2038-39",  # Year 25 typo from the image!
        "2040-41"   # Year 26 typo from the image!
    ]

    for idx, (yn, ye) in enumerate(zip(years_num, years_ending)):
        col_idx = 4 + idx  # Starts at D
        col_letter = get_column_letter(col_idx)
        
        ws_dep[f"{col_letter}3"] = yn
        ws_dep[f"{col_letter}3"].font = font_header_bold
        ws_dep[f"{col_letter}3"].alignment = align_center
        ws_dep[f"{col_letter}3"].border = border_dark_bottom

        ws_dep[f"{col_letter}4"] = ye
        ws_dep[f"{col_letter}4"].font = font_header_bold
        ws_dep[f"{col_letter}4"].alignment = align_center
        ws_dep[f"{col_letter}4"].border = border_dark_bottom

    ws_dep.row_dimensions[3].height = 20
    ws_dep.row_dimensions[4].height = 20

    # Total depreciable cost
    ws_dep["A6"] = "Total depreciable cost"
    ws_dep["A6"].font = font_regular
    ws_dep["C6"] = "=C30"
    ws_dep["C6"].font = font_bold
    ws_dep["C6"].border = border_all_thin
    ws_dep["C6"].number_format = fmt_decimal

    # Depreciable Life
    ws_dep["A7"] = "Depreciable Life"
    ws_dep["A7"].font = font_regular
    ws_dep["B7"] = 25.00
    ws_dep["B7"].font = font_regular
    ws_dep["B7"].number_format = "0.00"
    ws_dep["C7"] = "Years"
    ws_dep["C7"].font = font_regular

    # Salvage Value
    ws_dep["A8"] = "Salvage Value"
    ws_dep["A8"].font = font_regular
    ws_dep["B8"] = 0.10
    ws_dep["B8"].font = font_regular
    ws_dep["B8"].number_format = fmt_rate

    # Book Depreciation Section
    ws_dep["A10"] = "After 10 years, balance asset value equally spread"
    ws_dep["A10"].font = font_bold_italic

    ws_dep["A11"] = "Depreciation in Books"
    ws_dep["A11"].font = font_bold

    # Book Depreciation calculations
    for idx in range(26):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)
        
        if idx == 0:
            # Year 1 (1 month)
            formula = "=ROUND((C28*0.02 + C29*0.08)*1/12, 2)"
        elif idx < 10:
            # Year 2 to 10
            formula = "=ROUND(($C$28*0.02 + $C$29*0.08), 2)"
        else:
            # Year 11 to 26
            formula = "=ROUND(($C$6 - SUM($D$11:$M$11))/15, 2)"
            
        ws_dep[f"{col_letter}11"] = formula
        ws_dep[f"{col_letter}11"].font = font_regular
        ws_dep[f"{col_letter}11"].alignment = align_right
        ws_dep[f"{col_letter}11"].number_format = fmt_decimal
        ws_dep[f"{col_letter}11"].border = border_all_thin

    # IT WDV Depreciation Section
    ws_dep["A13"] = "As Per Income Tax - WDV"
    ws_dep["A13"].font = font_bold

    # Civil Works 10%
    ws_dep["A15"] = "Civil Works"
    ws_dep["A15"].font = font_bold
    ws_dep["C15"] = 0.10
    ws_dep["C15"].font = font_bold
    ws_dep["C15"].number_format = fmt_rate
    ws_dep["C15"].border = border_all_thin

    ws_dep["B16"] = "Opening Value"
    ws_dep["B17"] = "Depreciation"
    ws_dep["B18"] = "Closing Value"

    for r in [16, 17, 18]:
        ws_dep[f"B{r}"].font = font_regular

    for idx in range(26):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)
        
        # Opening Value
        if idx == 0:
            ws_dep[f"{col_letter}16"] = "=C28"
        else:
            prev_col = get_column_letter(col_idx - 1)
            ws_dep[f"{col_letter}16"] = f"={prev_col}18"
            
        # Depreciation
        if idx == 0:
            ws_dep[f"{col_letter}17"] = f"=ROUND({col_letter}16 * $C$15 * 9/12, 2)"
        else:
            ws_dep[f"{col_letter}17"] = f"=ROUND({col_letter}16 * $C$15, 2)"
            
        # Closing Value
        if idx == 2:  # Year 3 Closing Value typo override
            ws_dep[f"{col_letter}18"] = 1.02
        else:
            ws_dep[f"{col_letter}18"] = f"=ROUND({col_letter}16 - {col_letter}17, 2)"

        for r in [16, 17, 18]:
            cell = ws_dep[f"{col_letter}{r}"]
            cell.font = font_regular
            cell.alignment = align_right
            cell.number_format = fmt_decimal
            cell.border = border_all_thin

    # Plant & Machinery 15%
    ws_dep["A20"] = "Plant & Machinery"
    ws_dep["A20"].font = font_bold
    ws_dep["C20"] = 0.15
    ws_dep["C20"].font = font_bold
    ws_dep["C20"].number_format = fmt_rate
    ws_dep["C20"].border = border_all_thin

    ws_dep["B21"] = "Opening Value"
    ws_dep["B22"] = "Depreciation"
    ws_dep["B23"] = "Closing Value"

    for r in [21, 22, 23]:
        ws_dep[f"B{r}"].font = font_regular

    for idx in range(26):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)
        
        # Opening Value
        if idx == 0:
            ws_dep[f"{col_letter}21"] = "=C29"
        else:
            prev_col = get_column_letter(col_idx - 1)
            ws_dep[f"{col_letter}21"] = f"={prev_col}23"
            
        # Depreciation
        if idx == 0:
            ws_dep[f"{col_letter}22"] = f"=ROUND({col_letter}21 * $C$20 * 9/12, 2)"
        elif idx == 2:  # Year 3 Depreciation typo override
            ws_dep[f"{col_letter}22"] = 0.20
        else:
            ws_dep[f"{col_letter}22"] = f"=ROUND({col_letter}21 * $C$20, 2)"
            
        # Closing Value
        ws_dep[f"{col_letter}23"] = f"=ROUND({col_letter}21 - {col_letter}22, 2)"

        for r in [21, 22, 23]:
            cell = ws_dep[f"{col_letter}{r}"]
            cell.font = font_regular
            cell.alignment = align_right
            cell.number_format = fmt_decimal
            cell.border = border_all_thin

    # Total IT Depreciation
    ws_dep["A25"] = "Total IT Depreciation"
    ws_dep["A25"].font = font_bold
    
    for idx in range(26):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)
        ws_dep[f"{col_letter}25"] = f"={col_letter}17+{col_letter}22"
        ws_dep[f"{col_letter}25"].font = font_bold
        ws_dep[f"{col_letter}25"].alignment = align_right
        ws_dep[f"{col_letter}25"].number_format = fmt_decimal
        ws_dep[f"{col_letter}25"].border = border_bottom_double

    # Bottom Table: Capitalization
    ws_dep["A27"] = "Component"
    ws_dep["A27"].font = font_header_bold
    ws_dep["C27"] = "Capitalized Amount in Rs."
    ws_dep["C27"].font = font_header_bold
    
    for r in [27, 28, 29, 30, 31]:
        ws_dep[f"A{r}"].border = border_all_thin
        ws_dep[f"C{r}"].border = border_all_thin

    ws_dep["A28"] = "Civil Works"
    ws_dep["A28"].font = font_regular
    ws_dep["C28"] = 1.37
    ws_dep["C28"].font = font_regular
    ws_dep["C28"].number_format = fmt_decimal

    ws_dep["A29"] = "Plant & Machinery"
    ws_dep["A29"].font = font_regular
    ws_dep["C29"] = 7.00
    ws_dep["C29"].font = font_regular
    ws_dep["C29"].number_format = fmt_decimal

    ws_dep["A30"] = "Depreciable Base"
    ws_dep["A30"].font = font_bold
    ws_dep["C30"] = "=SUM(C28:C29)"
    ws_dep["C30"].font = font_bold
    ws_dep["C30"].number_format = fmt_decimal

    ws_dep["A31"] = "Weighted Average Rate of Depreciation"
    ws_dep["A31"].font = font_italic
    ws_dep["C31"] = "=ROUND((C28*C15+C29*C20)/C30, 4)"
    ws_dep["C31"].font = font_italic
    ws_dep["C31"].number_format = fmt_percent

    # IT Rates Table
    ws_dep["A33"] = "Depreciation Rates as per Income Tax"
    ws_dep["A33"].font = font_bold

    ws_dep["A34"] = "Under IT Act (WDV) - Equipment"
    ws_dep["A34"].font = font_regular
    ws_dep["C34"] = 0.15
    ws_dep["C34"].font = font_regular
    ws_dep["C34"].number_format = fmt_percent

    ws_dep["A35"] = "Under IT Act (WDV) - Buildings"
    ws_dep["A35"].font = font_regular
    ws_dep["C35"] = 0.10
    ws_dep["C35"].font = font_regular
    ws_dep["C35"].number_format = fmt_percent

    for r in [34, 35]:
        ws_dep[f"A{r}"].border = border_all_thin
        ws_dep[f"C{r}"].border = border_all_thin

    # Set column widths for Depreciation sheet
    ws_dep.column_dimensions["A"].width = 38
    ws_dep.column_dimensions["B"].width = 18
    ws_dep.column_dimensions["C"].width = 25
    for col_idx in range(4, 30):
        ws_dep.column_dimensions[get_column_letter(col_idx)].width = 11

    # ----------------------------------------------------
    # POPULATE FINANCIAL ANALYSIS SHEET (MAIN SHEET)
    # ----------------------------------------------------
    # Title Row
    ws.merge_cells("A1:AC1")
    ws["A1"] = "Financial Analysis"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws["A1"].fill = PatternFill(start_color=color_title_bg, end_color=color_title_bg, fill_type="solid")
    ws.row_dimensions[1].height = 35

    # Main Sheet Header
    ws.merge_cells("B2:C2")
    ws["B2"] = "Main Sheet"
    ws["B2"].font = font_main_sheet
    ws["B2"].alignment = align_center
    ws["B2"].fill = PatternFill(start_color=color_main_sheet_bg, end_color=color_main_sheet_bg, fill_type="solid")
    ws.row_dimensions[2].height = 20

    # Headers: Year and Year March Ending
    ws["A4"] = "Year"
    ws["A4"].font = font_header_bold
    ws["A4"].alignment = align_left

    ws["A5"] = "Year March Ending"
    ws["A5"].font = font_header_bold
    ws["A5"].alignment = align_left

    for idx, (yn, ye) in enumerate(zip(years_num, years_ending)):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)
        
        ws[f"{col_letter}4"] = yn
        ws[f"{col_letter}4"].font = font_header_bold
        ws[f"{col_letter}4"].alignment = align_center
        ws[f"{col_letter}4"].border = border_dark_bottom

        ws[f"{col_letter}5"] = ye
        ws[f"{col_letter}5"].font = font_header_bold
        ws[f"{col_letter}5"].alignment = align_center
        ws[f"{col_letter}5"].border = border_dark_bottom

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    # DSCR Section Labels
    ws.merge_cells("A7:A9")
    ws["A7"] = "DSCR Calculation"
    ws["A7"].font = font_bold
    ws["A7"].alignment = align_vertical_center
    ws["A7"].fill = PatternFill(start_color=color_section_bg, end_color=color_section_bg, fill_type="solid")

    ws["B7"] = "Cash accrual (PAT+Dep+Interest)"
    ws["B8"] = "Debt repayment (Interest + Principal)"
    ws["B9"] = "DSCR"

    for r in [7, 8, 9]:
        ws[f"B{r}"].font = font_bold if r == 9 else font_regular
        ws[f"B{r}"].alignment = align_left

    # Debt Repayment values (26 columns)
    debt_repay_vals = [
        35.75, 94.36, 90.78, 87.21, 83.63, 80.06, 76.48, 72.90, 69.33, 65.75,
        62.18, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00, 0.00,
        0.00, 0.00, 0.00, 0.00, 0.00, 0.00
    ]

    for idx, dr in enumerate(debt_repay_vals):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)

        # Cash Accrual cell (Formula: PAT + Depreciation + Interest)
        ws[f"{col_letter}7"] = f"={col_letter}15+{col_letter}16+{col_letter}17"
        ws[f"{col_letter}7"].font = font_regular
        ws[f"{col_letter}7"].alignment = align_right
        ws[f"{col_letter}7"].number_format = fmt_decimal
        ws[f"{col_letter}7"].border = border_all_thin

        # Debt Repayment cell
        ws[f"{col_letter}8"] = dr
        ws[f"{col_letter}8"].font = font_regular
        ws[f"{col_letter}8"].alignment = align_right
        ws[f"{col_letter}8"].number_format = fmt_decimal
        ws[f"{col_letter}8"].border = border_all_thin

        # DSCR cell (Formula)
        ws[f"{col_letter}9"] = f"=IF({col_letter}8=0, 0.00, {col_letter}7/{col_letter}8)"
        ws[f"{col_letter}9"].font = font_bold
        ws[f"{col_letter}9"].alignment = align_right
        ws[f"{col_letter}9"].number_format = fmt_decimal
        ws[f"{col_letter}9"].border = border_bottom_double

    # Summary box for DSCR (MIN, Average, MAX)
    ws["B10"] = "MIN DSCR"
    ws["C10"] = "=MIN(D9:N9)"  # Years 1-11 where debt is repaid
    ws["B11"] = "Average DSCR"
    ws["C11"] = "=AVERAGE(D9:N9)"
    ws["B12"] = "MAX DSCR"
    ws["C12"] = "=MAX(D9:N9)"

    for r in [10, 11, 12]:
        ws[f"B{r}"].font = font_bold
        ws[f"B{r}"].alignment = align_right
        ws[f"C{r}"].font = font_bold
        ws[f"C{r}"].alignment = align_right
        ws[f"C{r}"].number_format = fmt_decimal
        ws[f"C{r}"].border = border_all_thin

    # Project IRR Section Labels
    ws.merge_cells("A14:A20")
    ws["A14"] = "Project IRR"
    ws["A14"].font = font_bold
    ws["A14"].alignment = align_vertical_center
    ws["A14"].fill = PatternFill(start_color=color_section_bg, end_color=color_section_bg, fill_type="solid")

    ws["B14"] = "Capital Expenditure"
    ws["B15"] = "PAT"
    ws["B16"] = "Depreciation"
    ws["B17"] = "Interest"
    ws["B18"] = "Salvage Value"
    ws["B19"] = "Incremental Margin Money for WC"
    ws["B20"] = "Net Cash Flow"
    ws["B21"] = "Project IRR"

    ws["C14"] = -788.72
    ws["C14"].font = font_regular
    ws["C14"].alignment = align_right
    ws["C14"].number_format = fmt_decimal
    ws["C14"].border = border_all_thin

    ws["C20"] = 0.00
    ws["C20"].font = font_bold
    ws["C20"].alignment = align_right
    ws["C20"].number_format = fmt_decimal
    ws["C20"].border = border_all_thin

    # Values for Project IRR (26 columns)
    pat_vals = [
        24.09, 49.54, 41.30, 43.58, 45.62, 47.84, 50.02, 52.18, 54.30, 56.38,
        58.43, 60.44, 62.41, 64.33, 66.22, 68.05, 68.77, 67.67, 61.53, 60.40,
        59.21, 57.96, 56.65, 55.27, 73.77, 73.77
    ]
    interest_vals = [
        26.82, 52.29, 48.72, 45.14, 41.57, 37.99, 34.41, 30.84, 27.26, 23.69,
        20.11, 16.54, 12.96, 9.39, 5.80, 2.22, 0.00, 0.00, 0.00, 0.00,
        0.00, 0.00, 0.00, 0.00, 0.00, 0.00
    ]
    wc_vals = [
        2.41, 2.40, 0.58, 0.03, 0.02, 0.03, 0.03, 0.04, 0.04, 0.04,
        0.04, 0.04, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.05, 0.06,
        0.07, 0.07, 0.08, 0.00, 0.00, 0.00
    ]

    for idx in range(26):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)

        # Capital Expenditure (blank)
        ws[f"{col_letter}14"].font = font_regular
        ws[f"{col_letter}14"].border = border_all_thin

        # PAT
        ws[f"{col_letter}15"] = pat_vals[idx]
        ws[f"{col_letter}15"].font = font_regular
        ws[f"{col_letter}15"].alignment = align_right
        ws[f"{col_letter}15"].number_format = fmt_decimal
        ws[f"{col_letter}15"].border = border_all_thin

        # Depreciation (Formula linking to Depreciation sheet)
        ws[f"{col_letter}16"] = f"=Depreciation!{col_letter}11"
        ws[f"{col_letter}16"].font = font_regular
        ws[f"{col_letter}16"].alignment = align_right
        ws[f"{col_letter}16"].number_format = fmt_decimal
        ws[f"{col_letter}16"].border = border_all_thin

        # Interest
        ws[f"{col_letter}17"] = interest_vals[idx]
        ws[f"{col_letter}17"].font = font_regular
        ws[f"{col_letter}17"].alignment = align_right
        ws[f"{col_letter}17"].number_format = fmt_decimal
        ws[f"{col_letter}17"].border = border_all_thin

        # Salvage Value (78.87 in Year 26, index 25)
        if idx == 25:
            ws[f"{col_letter}18"] = 78.87
        else:
            ws[f"{col_letter}18"] = 0.00
        ws[f"{col_letter}18"].font = font_regular
        ws[f"{col_letter}18"].alignment = align_right
        ws[f"{col_letter}18"].number_format = fmt_decimal
        ws[f"{col_letter}18"].border = border_all_thin

        # WC
        ws[f"{col_letter}19"] = wc_vals[idx]
        ws[f"{col_letter}19"].font = font_regular
        ws[f"{col_letter}19"].alignment = align_right
        ws[f"{col_letter}19"].number_format = fmt_decimal
        ws[f"{col_letter}19"].border = border_all_thin

        # Net Cash Flow (Formula)
        if idx == 0:
            ws[f"{col_letter}20"] = f"={col_letter}14+{col_letter}15+{col_letter}16+{col_letter}17+{col_letter}18-{col_letter}19+C14"
        else:
            ws[f"{col_letter}20"] = f"={col_letter}14+{col_letter}15+{col_letter}16+{col_letter}17+{col_letter}18-{col_letter}19"
        
        ws[f"{col_letter}20"].font = font_bold
        ws[f"{col_letter}20"].alignment = align_right
        ws[f"{col_letter}20"].number_format = fmt_decimal
        ws[f"{col_letter}20"].border = border_bottom_double

    for r in range(14, 21):
        ws[f"B{r}"].font = font_bold if r == 20 else font_regular
        ws[f"B{r}"].alignment = align_left

    # Project IRR cell formula
    ws["B21"].font = font_bold
    ws["B21"].alignment = align_right
    ws["C21"] = "=IRR(D20:AC20)"
    ws["C21"].font = font_bold
    ws["C21"].alignment = align_right
    ws["C21"].number_format = fmt_percent
    ws["C21"].border = border_all_thin

    # Equity IRR Section Labels
    ws.merge_cells("A24:A29")
    ws["A24"] = "Equity IRR"
    ws["A24"].font = font_bold
    ws["A24"].alignment = align_vertical_center
    ws["A24"].fill = PatternFill(start_color=color_section_bg, end_color=color_section_bg, fill_type="solid")

    ws["B24"] = "Equity Investment"
    ws["B25"] = "PAT"
    ws["B26"] = "Depreciation"
    ws["B27"] = "Salvage Value"
    ws["B28"] = "Less Repayment of Term Loan"
    ws["B29"] = "Net Cash Flow"
    ws["B30"] = "Equity IRR"

    ws["C24"] = -157.74
    ws["C24"].font = font_regular
    ws["C24"].alignment = align_right
    ws["C24"].number_format = fmt_decimal
    ws["C24"].border = border_all_thin

    ws["C29"].border = border_all_thin

    # Term Loan Repayment (26 columns)
    term_loan_repay = [
        0.00, 42.07, 42.07, 42.07, 42.07, 42.07, 42.07, 42.07, 42.07, 42.07,
        42.07, 42.07, 42.07, 42.07, 42.07, 42.07, 0.00, 0.00, 0.00, 0.00,
        0.00, 0.00, 0.00, 0.00, 0.00, 0.00
    ]

    for idx in range(26):
        col_idx = 4 + idx
        col_letter = get_column_letter(col_idx)

        # Equity Investment (blank)
        ws[f"{col_letter}24"].font = font_regular
        ws[f"{col_letter}24"].border = border_all_thin

        # PAT (Formula referencing Row 15)
        ws[f"{col_letter}25"] = f"={col_letter}15"
        ws[f"{col_letter}25"].font = font_regular
        ws[f"{col_letter}25"].alignment = align_right
        ws[f"{col_letter}25"].number_format = fmt_decimal
        ws[f"{col_letter}25"].border = border_all_thin

        # Depreciation (Formula referencing Row 16)
        ws[f"{col_letter}26"] = f"={col_letter}16"
        ws[f"{col_letter}26"].font = font_regular
        ws[f"{col_letter}26"].alignment = align_right
        ws[f"{col_letter}26"].number_format = fmt_decimal
        ws[f"{col_letter}26"].border = border_all_thin

        # Salvage Value (78.87 in Year 26, index 25)
        if idx == 25:
            ws[f"{col_letter}27"] = 78.87
        else:
            ws[f"{col_letter}27"] = 0.00
        ws[f"{col_letter}27"].font = font_regular
        ws[f"{col_letter}27"].alignment = align_right
        ws[f"{col_letter}27"].number_format = fmt_decimal
        ws[f"{col_letter}27"].border = border_all_thin

        # Less Repayment of Term Loan
        ws[f"{col_letter}28"] = term_loan_repay[idx]
        ws[f"{col_letter}28"].font = font_regular
        ws[f"{col_letter}28"].alignment = align_right
        ws[f"{col_letter}28"].number_format = fmt_decimal
        ws[f"{col_letter}28"].border = border_all_thin

        # Net Cash Flow (Formula)
        if idx == 0:
            ws[f"{col_letter}29"] = f"={col_letter}25+{col_letter}26+{col_letter}27-{col_letter}28+C24"
        else:
            ws[f"{col_letter}29"] = f"={col_letter}25+{col_letter}26+{col_letter}27-{col_letter}28"
        
        ws[f"{col_letter}29"].font = font_bold
        ws[f"{col_letter}29"].alignment = align_right
        ws[f"{col_letter}29"].number_format = fmt_decimal
        ws[f"{col_letter}29"].border = border_bottom_double

    for r in range(24, 30):
        ws[f"B{r}"].font = font_bold if r == 29 else font_regular
        ws[f"B{r}"].alignment = align_left

    # Equity IRR cell formula
    ws["B30"].font = font_bold
    ws["B30"].alignment = align_right
    ws["C30"] = "=IRR(D29:AC29)"
    ws["C30"].font = font_bold
    ws["C30"].alignment = align_right
    ws["C30"].number_format = fmt_percent
    ws["C30"].border = border_all_thin

    # Set row heights
    for r in range(14, 22):
        ws.row_dimensions[r].height = 20
    for r in range(24, 31):
        ws.row_dimensions[r].height = 20

    # Column Width Formatting
    ws.column_dimensions["A"].width = 20  # Section label
    ws.column_dimensions["B"].width = 38  # Row description
    ws.column_dimensions["C"].width = 15  # Investment column
    
    # Auto-fit year columns
    for col_idx in range(4, 30):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 11

    # Save to file
    output_path = "c:/Users/Asus 15 Aspire/Desktop/another/Financial_Analysis_Model.xlsx"
    wb.save(output_path)
    print(f"Excel file generated successfully at: {output_path}")

if __name__ == "__main__":
    create_financial_model()

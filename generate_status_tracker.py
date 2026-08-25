import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import CellIsRule

def create_status_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Status Tracker"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Styling helper tokens
    FONT_NAME = "Segoe UI"
    font_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_NAME, size=11)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Deep Navy Blue
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Set Column Widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    # Headers
    ws["A1"] = "Date"
    ws["B1"] = "Status"
    
    for col in ["A1", "B1"]:
        cell = ws[col]
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
    
    # Data rows to insert (matching the screenshot and a few more examples)
    data = [
        ("2025-04-15", None), # 15/Apr/2025 (Expired)
        ("2027-03-26", None), # 26/Mar/2027 (Sufficient Time)
        ("2026-06-29", None), # 29/Jun/2026 (Expiring Soon - 6 days away)
        ("2026-07-23", None), # 23/Jul/2026 (Sufficient Time - 30 days away)
        (None, None),         # Empty Row
        ("2026-06-22", None), # 22/Jun/2026 (Expired - Yesterday)
        ("2026-06-24", None), # 24/Jun/2026 (Expiring Soon - Tomorrow)
        ("2026-06-30", None), # 30/Jun/2026 (Sufficient Time - 7 days away)
    ]

    # Date formatting string: DD/MMM/YYYY
    date_format = "dd/mmm/yyyy"

    for idx, (date_val, _) in enumerate(data):
        row = idx + 2
        
        # Write Date
        cell_date = ws.cell(row=row, column=1)
        if date_val:
            # Parse date string to write it as an actual Excel date
            import datetime
            dt = datetime.datetime.strptime(date_val, "%Y-%m-%d").date()
            cell_date.value = dt
            cell_date.number_format = date_format
        else:
            cell_date.value = ""
            
        cell_date.font = font_body
        cell_date.alignment = align_center
        cell_date.border = border_cell

        # Write Status Formula
        cell_status = ws.cell(row=row, column=2)
        # Using nested IF instead of IFS for maximum compatibility
        cell_status.value = f'=IF(A{row}="", "", IF((A{row}-TODAY())<0, "Expired", IF((A{row}-TODAY())<7, "Expiring Soon", "Sufficient Time")))'
        cell_status.font = font_body
        cell_status.alignment = align_center
        cell_status.border = border_cell

    # Apply conditional formatting for colors in Column B (B2:B50)
    # 1. Expired (Soft Red Fill)
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    red_font = Font(name=FONT_NAME, size=11, bold=True, color="C00000")
    # 2. Expiring Soon (Soft Yellow Fill)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    yellow_font = Font(name=FONT_NAME, size=11, bold=True, color="7F6000")
    # 3. Sufficient Time (Soft Green Fill)
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    green_font = Font(name=FONT_NAME, size=11, bold=True, color="385723")

    ws.conditional_formatting.add("B2:B100", CellIsRule(operator="equal", formula=['"Expired"'], stopIfTrue=True, fill=red_fill, font=red_font))
    ws.conditional_formatting.add("B2:B100", CellIsRule(operator="equal", formula=['"Expiring Soon"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
    ws.conditional_formatting.add("B2:B100", CellIsRule(operator="equal", formula=['"Sufficient Time"'], stopIfTrue=True, fill=green_fill, font=green_font))

    # Save to file
    filename = "Date_Status_Tracker_v2.xlsx"
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    print(f"Excel file created successfully at: {filepath}")

if __name__ == "__main__":
    create_status_tracker()

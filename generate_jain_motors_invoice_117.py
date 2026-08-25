import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def create_jain_motors_invoice_117():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Challan 117"

    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Fonts
    FONT_NAME = "Segoe UI"
    font_store_name = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
    font_store_sub = Font(name=FONT_NAME, size=10, italic=True, color="F8F9FA")
    font_section_lbl = Font(name=FONT_NAME, size=10, bold=True, color="495057")
    font_section_val = Font(name=FONT_NAME, size=10, color="212529")
    font_header = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_NAME, size=10, color="212529")
    font_body_bold = Font(name=FONT_NAME, size=10, bold=True, color="212529")
    font_footer = Font(name=FONT_NAME, size=9, italic=True, color="6C757D")
    font_sig_title = Font(name=FONT_NAME, size=9, bold=True, color="495057")
    font_paid_badge = Font(name=FONT_NAME, size=14, bold=True, color="385723")

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Borders
    border_color_gray = "D9D9D9"
    border_color_dark = "212529"
    
    side_thin = Side(border_style="thin", color=border_color_gray)
    side_thick_bottom = Side(border_style="medium", color=border_color_dark)
    side_double = Side(border_style="double", color=border_color_dark)
    side_top_thin = Side(border_style="thin", color=border_color_dark)
    
    border_cell = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    border_header = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thick_bottom)
    border_summary = Border(top=side_top_thin, bottom=side_double)
    border_paid = Border(left=Side(border_style="medium", color="385723"),
                         right=Side(border_style="medium", color="385723"),
                         top=Side(border_style="medium", color="385723"),
                         bottom=Side(border_style="medium", color="385723"))

    # Color Fills
    fill_store_header = PatternFill(start_color="212529", end_color="212529", fill_type="solid") # Dark Charcoal
    fill_table_header = PatternFill(start_color="495057", end_color="495057", fill_type="solid") # Slate Gray
    fill_zebra_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_zebra_odd = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    fill_summary = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")       # Light Gray for totals
    fill_paid_bg = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")       # Soft Green

    # Helper function to style ranges
    def style_range(ws, cell_range, font=None, alignment=None, fill=None, border=None):
        for row in ws[cell_range]:
            for cell in row:
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border

    # Set column widths
    column_widths = {
        "A": 3,   # Left margin
        "B": 8,   # S.No.
        "C": 35,  # Particulars
        "D": 10,  # Qty
        "E": 14,  # Rate
        "F": 18,  # Amount
        "G": 3    # Right margin
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row Heights
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 12
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 15

    # 1. Main Header Title
    ws.merge_cells("B2:F2")
    ws["B2"] = "JAIN MOTORS"
    style_range(ws, "B2:F2", font=font_store_name, alignment=align_center, fill=fill_store_header)

    ws.merge_cells("B3:F3")
    ws["B3"] = "CHALLAN BOOK  |  Mob. : 9358993882"
    style_range(ws, "B3:F3", font=font_store_sub, alignment=align_center, fill=fill_store_header)

    # 2. Metadata details (Challan No, Date, Customer)
    # Row 5
    ws["B5"] = "Challan No:"
    ws["B5"].font = font_section_lbl
    ws["B5"].alignment = align_left
    
    ws["C5"] = 117
    ws["C5"].font = font_section_val
    ws["C5"].alignment = align_left
    
    ws["E5"] = "Date:"
    ws["E5"].font = font_section_lbl
    ws["E5"].alignment = align_right
    
    ws["F5"] = "2026-06-08" # June 8, 2026
    ws["F5"].font = font_section_val
    ws["F5"].alignment = align_left
    
    # Row 6
    ws["B6"] = "Customer:"
    ws["B6"].font = font_section_lbl
    ws["B6"].alignment = align_left
    
    ws["C6"] = "MSC."
    ws["C6"].font = font_section_val
    ws["C6"].alignment = align_left

    # Draw border around the metadata block
    style_range(ws, "B5:F6", border=Border(bottom=side_thin))

    # 3. Main Table Headers (Row 8)
    headers = [
        ("S.No.", align_center),
        ("Particulars", align_left),
        ("Qty", align_center),
        ("Rate", align_right),
        ("Amount", align_right)
    ]
    
    ws.row_dimensions[8].height = 26
    for idx, (text, align) in enumerate(headers, start=2):
        cell = ws.cell(row=8, column=idx, value=text)
        cell.font = font_header
        cell.alignment = align
        cell.fill = fill_table_header
        cell.border = border_header

    # 4. Transcribed Data (Rows 9 to 12)
    items_data = [
        ["ROUND BELTS X 4", 4, 1550.00],
        ["ROUND BELT 10 X 8", 8, 5675.00],
        ["ROUND BELT 20 X 8", 4, 11360.00],
        ["SAFTEY BELT", 4, 2000.00]
    ]

    for i, data in enumerate(items_data):
        row_idx = 9 + i
        ws.row_dimensions[row_idx].height = 22
        fill_zebra = fill_zebra_odd if row_idx % 2 == 1 else fill_zebra_even
        
        # S.No.
        ws.cell(row=row_idx, column=2, value=i+1).alignment = align_center
        # Particulars
        ws.cell(row=row_idx, column=3, value=data[0]).alignment = align_left
        # Qty
        ws.cell(row=row_idx, column=4, value=data[1]).alignment = align_center
        ws.cell(row=row_idx, column=4).number_format = "#,##0"
        # Rate
        ws.cell(row=row_idx, column=5, value=data[2]).alignment = align_right
        ws.cell(row=row_idx, column=5).number_format = "₹#,##0.00"
        # Amount (Formula)
        amt_cell = ws.cell(row=row_idx, column=6, value=f"=D{row_idx}*E{row_idx}")
        amt_cell.alignment = align_right
        amt_cell.number_format = "₹#,##0.00"
        
        # Apply borders, font and fill
        for col_idx in range(2, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = border_cell
            cell.fill = fill_zebra

    # 5. Totals Block (Row 13)
    ws.row_dimensions[13].height = 24

    # Grand Total (Calculated via formula)
    ws.merge_cells("B13:E13")
    ws["B13"] = "Total"
    style_range(ws, "B13:E13", font=font_body_bold, alignment=align_right, fill=fill_summary, border=border_summary)
    
    total_cell = ws.cell(row=13, column=6, value="=SUM(F9:F12)")
    total_cell.font = Font(name=FONT_NAME, size=11, bold=True, color="212529")
    total_cell.alignment = align_right
    total_cell.number_format = "₹#,##0.00"
    total_cell.fill = fill_summary
    total_cell.border = border_summary

    # Spacer row
    ws.row_dimensions[14].height = 15

    # 6. Paid Badge (Row 15 & 16)
    # The handwritten note has "PAID" on it. Let's make a beautiful stamp.
    ws.merge_cells("B15:C16")
    ws["B15"] = "PAID"
    style_range(ws, "B15:C16", font=font_paid_badge, alignment=align_center, fill=fill_paid_bg, border=border_paid)

    # 7. Brand Footers Block
    ws.merge_cells("B18:F18")
    ws["B18"] = "Authorized Dealer: FERRETERRO | BOSCH | TAPARIA | INGCO"
    style_range(ws, "B18:F18", font=font_footer, alignment=align_center)

    ws.row_dimensions[19].height = 10
    ws.row_dimensions[20].height = 20
    ws.row_dimensions[21].height = 15
    ws.row_dimensions[22].height = 15
    ws.row_dimensions[23].height = 15

    # Signature Block (Bottom Right)
    ws.merge_cells("E20:F20")
    ws["E20"] = "For JAIN MOTORS"
    style_range(ws, "E20:F20", font=font_sig_title, alignment=align_center)

    ws.merge_cells("E22:F22")
    ws["E22"] = "_________________________"
    style_range(ws, "E22:F22", font=font_body, alignment=align_center)

    ws.merge_cells("E23:F23")
    ws["E23"] = "Authorized Signature"
    style_range(ws, "E23:F23", font=font_footer, alignment=align_center)

    # Save Excel Workbook
    output_filename = "Jain_Motors_Challan_117.xlsx"
    wb.save(output_filename)
    print(f"Jain Motors Challan 117 created successfully: {output_filename}")

if __name__ == "__main__":
    create_jain_motors_invoice_117()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_exact_replica():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KHAVDA 442MW"
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    
    # ---------------- STYLES ----------------
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    font_header_bold = Font(name=font_family, size=10, bold=True)
    font_data = Font(name=font_family, size=9.5)
    font_data_bold = Font(name=font_family, size=9.5, bold=True)
    font_total = Font(name=font_family, size=11, bold=True, color="1F4E78")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_border_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

    num_fmt = '#,##,##0.00'
    num_fmt_int = '#,##,##0'

    # ---------------- ROW 1: TOP SUMMARY HEADER ----------------
    ws.cell(row=1, column=8, value=72768871.00).number_format = num_fmt
    ws.cell(row=1, column=8).font = font_header_bold
    ws.cell(row=1, column=8).fill = yellow_fill
    ws.cell(row=1, column=8).alignment = align_right
    ws.cell(row=1, column=8).border = thin_border
    
    ws.cell(row=1, column=9, value=14130554.06).number_format = num_fmt
    ws.cell(row=1, column=9).font = font_header_bold
    ws.cell(row=1, column=9).fill = yellow_fill
    ws.cell(row=1, column=9).alignment = align_right
    ws.cell(row=1, column=9).border = thin_border
    ws.row_dimensions[1].height = 20

    # ---------------- ROW 2: TABLE HEADERS ----------------
    headers = [
        "SR NO.", "PO NUMBER", "Project Name", "Type", 
        "Description", "Company Name", "VENDOR NAME", 
        "Paid As On", "Amount to be paid"
    ]
    
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.font = font_header_bold
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[2].height = 25

    # ---------------- ROWS 3 to 46: TABLE DATA ----------------
    raw_data = [
        (8, 6002526032, "KHAVDA 442MW", "Service", "Piling Work", "VIBRANT CONSTRUCTION COMPANY", "Mohamad (6360589955)", 8142757, 500000, False, False),
        (16, 6002526062, "KHAVDA 442MW", "Service", "Erection of MMS", "Murlidhar Fabrication", "Suresh Kumar (+91 99783 01708)", 2684889, 1460059, False, False),
        (17, 6002526026, "KHAVDA 442MW", "Service", "Erection of MMS", "Choudhary Electric Constrution", "Sarvan Choudhary +91 90013 29239", 8361290, 2507431, False, False),
        (22, 6002526025, "KHAVDA 442MW", "Service", "Supply Of White Sand", "R.K.PROJECT", "GAGAL RAMESHBHAI (+91 6354263514)", 5216596, 24170, False, False),
        (24, 6002526050, "KHAVDA 442MW", "Service", "DG", "PALIWAL BUILDING MATERIAL", "Lalit-", 1606082, 250000, False, False),
        (25, 6002526051, "KHAVDA 442MW", "Service", "HT Cable laying", "SUPER ELECTRO POWER POONA RAM", "Poona ram", 3817310, 213861, False, False),
        (27, 6002526061, "KHAVDA 442MW", "Service", "VLF Testing", "TNM GENIE PRIVATE LIMITED", "Saravanakkumar +91 9741402787", 664070, 44314, True, False), # Yellow Comp
        (31, 6002526072, "KHAVDA 442MW", "Service", "Erection of MMS", "PUSHPA ENTERPRISES", "RAM KUMAR (+91 97259 87924)", 2426646, 222850, False, False),
        (15, 6002526084, "KHAVDA 442MW", "Service", "Erection of MMS", "R P CONSTRUCTION", "BAYUNANDAN PANDE (92419 50137)", 1651357, 35435, False, False),
        (17, 6002526087, "KHAVDA 442MW", "Service", "Erection of MMS", "AKSHAR CORPORATION", "MEHULKUMAR VASUDEVBHAI PATEL", 2636044, 456175, False, False),
        (19, 6002526091, "KHAVDA 442MW", "Service", "Erection of MMS", "ADITYA POWER TECHNOLOGIES", "Hemang Patel", 620240, 428614, False, False),
        (20, 6002526092, "KHAVDA 442MW", "Service", "Erection of MMS", "SABNAM ENTERPRISE", "SABNAM BANU", 2883657, 423482, False, False),
        (27, 6002526110, "KHAVDA 442MW", "Service", "Erection of MMS", "Polyvision India", "Mahesh Ganesh Patel (9819343954)", 1156520, 300000, False, False),
        (29, 6002526112, "KHAVDA 442MW", "Service", "Erection of MMS", "M/S NITYA DEV ENTERPRISES", "SANDEEP KUMAR (+91 9992279349, 9467496874)", 236952, 209366, False, False),
        (30, 6002526113, "KHAVDA 442MW", "Service", "AC-DC Work", "TD ELECTRICAL & SOLAR", "TAPOMOY DEY: +91 7602143142", 2280000, 150000, False, False),
        (31, 6002526115, "KHAVDA 442MW", "Service", "Erection of MMS", "Mahadev Infratech Enterprises", "Hari Ram 9983414884, 8239971600", 555841, 100361, False, False),
        (32, 6002526120, "KHAVDA 442MW", "Service", "LT & Earthing Trench Work", "R.K.PROJECT", "GAGAL RAMESHBHAI (+91 6354263514)", None, 608934, False, False),
        (36, 6002526132, "KHAVDA 442MW", "Service", "Hiring of Vechile", "R.S.B.T. SUPPLIERS", "Praveen bhalothia (+91 6005289994, 9879393598)", 1458428, 496440, False, False),
        (41, 6002526145, "KHAVDA 442MW", "Service", "Erection of MMS", "OZONE INFRA", "AMIT KUMAR BAJPAI (+91 9554523222)", 1635481, 908870, False, False),
        (43, 6002526152, "KHAVDA 442MW", "Service", "Erection of MMS", "TANSUMAN LIGHTNING ENTERPRISES (OPC) PRIVATE LIMITED", "Vishal Rathore (+91 7790909130)", 446072, 111519, False, False),
        (51, "", "KHAVDA 442MW", "Supplier", "Hardware & Machine", "Rajkot Machinery", "Rajkot Machinery", 1100000, 130023, False, False),
        (57, "", "KHAVDA 442MW", "Service", "Hiring Splicing Service", "New Radiant Enterprises", "New Radiant Enterprises", None, 61950, False, False),
        (5, 6002526034, "KHAVDA 442MW", "Service", "IDT Civil", "Dayachand -7", "PAWAN KUMAR 80555 27414", 2740748, 152263, True, True), # Yellow Comp & Paid
        (7, 6002526064, "KHAVDA 442MW", "Service", "IDT Civil", "Baburam-3", "BABU RAM (+91 9001678721)", 1160379, 128931, False, False),
        (20, 6002526001, "KHAVDA 442MW", "Service", "AC-DC Work", "J.P. ENTERPRISES", "JAI PRAKASH +91 98282 39611", 3535000, 200000, False, False),
        (42, 6002526027, "KHAVDA 442MW", "Service", "HT Jointer", "Sanjay Kumar Yadav", "Sanjay Kumar Yadav", 700000, 42306, False, False),
        (50, 6002526086, "KHAVDA 442MW", "Service", "Erection of MMS", "SARTHAK ENTERPRISES", "OM PRAKASH +91 8950588873, 9896098873", 17118, 4280, False, False),
        (77, 6002526136, "KHAVDA 442MW", "Service", "AC-DC Work", "FATIMA CONSTRUCTION", "ZAINAB FATMA 7428270013", 2200000, 450000, False, False),
        (91, 6002526155, "KHAVDA 442MW", "Service", "String Work", "FATA RAM", "FATA RAM (+91 8401688194)", 380700, 42300, False, False),
        (95, 6002526159, "KHAVDA 442MW", "Service", "Erection of Module", "Juhi Gulab", "ALHANOOR (+91 8460311858)", 329386, 93056, True, True), # Yellow Comp & Paid
        (135, "", "KHAVDA 442MW", "Service", "AC-DC Work", "Nagnararai Construction Company-2", "Lun Singh ( 9680881351)", 3220000, 700000, False, False),
        (140, "", "KHAVDA 442MW", "Service", "HT Cable laying", "Ashok Kumar Bairwa", "Ashok Kumar Bairwa(9740523254)", 350248, 25000, False, False),
        (143, "", "KHAVDA 442MW", "Service", "Supply Labour", "Girdhari Ram Labour", "Girdhari Ram Labour", 821512, 300000, False, False),
        (147, "", "KHAVDA 442MW", "Service", "DC Works", "Sumer Singh", "Devi Singh/ Sumer Singh", 1961000, 700000, False, False),
        (154, "", "KHAVDA 442MW", "Service", "AC-DC Work", "Jeeshan Ahemad", "Jeeshan Ahemad", 1625000, 100000, False, False),
        (157, "", "KHAVDA 442MW", "Service", "Hiring of Vechile", "SHARDA MAHALA", "RJ37-GB-0907", 73548, 52903, False, False),
        (158, "", "KHAVDA 442MW", "Service", "AC Work", "Abhishek Pandey", "Abhishek Pandey", 1495000, 200000, False, False),
        (159, "", "KHAVDA 442MW", "Service", "AC-DC Work", "ISHAN", "Sajjirul", 2100000, 260000, False, False),
        (173, "", "KHAVDA 442MW", "Service", "WATER SUPPLIERS", "Abdul Kayum Ibrahim Sama", "Abdul Kayum Ibrahim Sama", 300000, 170000, False, False),
        (177, "", "KHAVDA 442MW", "Service", "Erection of MMS", "JAHANGIR ALI", "JAHANGIR ALI", 179000, 45861, False, False),
        (182, "", "KHAVDA 442MW", "Service", "Oger Machine", "Padam Singh", "Padam SIngh", None, 39800, False, False),
        (185, "", "KHAVDA 442MW", "Service", "Hiring of Vechile", "Mohan Lal", "Mohan Lal", None, 40000, False, False),
        (186, 6002627006, "KHAVDA 442MW", "Service", "Hiring of Vechile", "TAJMAMAD LATIF SAMA", "TAJMAMAD LATIF SAMA", None, 40000, False, False),
        ("", "", "", "", "", "Gujrat Pota Cabin", "", None, 700000, False, False),
    ]

    for idx, row in enumerate(raw_data, 3):
        orig_sr, po_num, prj, vtype, desc, comp, vendor_name, paid, payable, y_comp, y_paid = row
        
        # Col 1: SR NO.
        ws.cell(row=idx, column=1, value=orig_sr).alignment = align_center
        
        # Col 2: PO NUMBER
        ws.cell(row=idx, column=2, value=str(po_num) if po_num else "").alignment = align_center
        
        # Col 3: Project Name
        ws.cell(row=idx, column=3, value=prj).alignment = align_center
        
        # Col 4: Type
        ws.cell(row=idx, column=4, value=vtype).alignment = align_center
        
        # Col 5: Description
        ws.cell(row=idx, column=5, value=desc).alignment = align_left
        
        # Col 6: Company Name
        c_comp = ws.cell(row=idx, column=6, value=comp)
        c_comp.alignment = align_left
        if y_comp:
            c_comp.fill = yellow_fill
            c_comp.font = font_data_bold
            
        # Col 7: VENDOR NAME
        ws.cell(row=idx, column=7, value=vendor_name).alignment = align_left
        
        # Col 8: Paid As On
        c_paid = ws.cell(row=idx, column=8, value=paid if paid is not None else "")
        if paid is not None:
            c_paid.number_format = num_fmt_int
            c_paid.alignment = align_right
        else:
            c_paid.alignment = align_center
        if y_paid:
            c_paid.fill = yellow_fill
            c_paid.font = font_data_bold
            
        # Col 9: Amount to be paid
        c_payable = ws.cell(row=idx, column=9, value=payable if payable is not None else "")
        if payable is not None:
            c_payable.number_format = num_fmt_int
            c_payable.alignment = align_right
            
        # Borders & Fonts for row
        for col_c in range(1, 10):
            cell = ws.cell(row=idx, column=col_c)
            if not cell.font or not getattr(cell.font, 'bold', False):
                cell.font = font_data
            cell.border = thin_border
            
        ws.row_dimensions[idx].height = 20

    # ---------------- ROW 47: GRAND TOTALS ----------------
    tot_row = len(raw_data) + 3
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=7)
    ws.cell(row=tot_row, column=1, value="TOTAL").font = font_total
    ws.cell(row=tot_row, column=1).alignment = align_right
    
    # Total Paid
    c_t_paid = ws.cell(row=tot_row, column=8, value=f"=SUM(H3:H{tot_row-1})")
    c_t_paid.font = font_total
    c_t_paid.number_format = num_fmt
    c_t_paid.alignment = align_right
    
    # Total Payable
    c_t_pay = ws.cell(row=tot_row, column=9, value=f"=SUM(I3:I{tot_row-1})")
    c_t_pay.font = font_total
    c_t_pay.number_format = num_fmt
    c_t_pay.alignment = align_right
    
    for c in range(1, 10):
        cell = ws.cell(row=tot_row, column=c)
        cell.fill = total_fill
        cell.border = double_bottom_border
        
    ws.row_dimensions[tot_row].height = 26

    # Column Widths matching original design layout
    ws.column_dimensions['A'].width = 8   # SR NO.
    ws.column_dimensions['B'].width = 16  # PO NUMBER
    ws.column_dimensions['C'].width = 16  # Project Name
    ws.column_dimensions['D'].width = 12  # Type
    ws.column_dimensions['E'].width = 24  # Description
    ws.column_dimensions['F'].width = 38  # Company Name
    ws.column_dimensions['G'].width = 36  # VENDOR NAME
    ws.column_dimensions['H'].width = 20  # Paid As On
    ws.column_dimensions['I'].width = 20  # Amount to be paid

    out_file = "KHAVDA_442MW_Statement_Exact_Replica.xlsx"
    wb.save(out_file)
    wb.save(f"c:\\Users\\Asus 15 Aspire\\Desktop\\{out_file}")
    print(f"Exact replica workbook saved successfully as: {out_file} on Desktop!")

if __name__ == "__main__":
    create_exact_replica()

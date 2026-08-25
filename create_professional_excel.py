import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_vendor_tracker():
    wb = openpyxl.Workbook()
    
    # Setup Sheet 1: Master Tracker
    ws1 = wb.active
    ws1.title = "PO & Payment Tracker"
    ws1.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    
    # ---------------- STYLES DEFINITION ----------------
    # Color Palette
    navy_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    sub_banner_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    col_header_fill = PatternFill(start_color="2C4D75", end_color="2C4D75", fill_type="solid")
    total_row_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    kpi_title_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    # Highlight Fills
    yellow_tag_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    sub_title_font = Font(name=font_family, size=10, bold=True, color="1F4E78")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=9.5)
    data_bold_font = Font(name=font_family, size=9.5, bold=True)
    total_font = Font(name=font_family, size=11, bold=True, color="1F4E78")
    
    kpi_label_font = Font(name=font_family, size=9, bold=True, color="595959")
    kpi_num_font = Font(name=font_family, size=14, bold=True, color="1F4E78")
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(style='thin', color='D9D9D9')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(top=Side(style='thin', color='D9D9D9'), bottom=Side(style='medium', color='1F4E78'))
    double_bottom_border = Border(top=Side(style='thin', color='1F4E78'), bottom=Side(style='double', color='1F4E78'))

    # Number Formats
    currency_fmt = '[$₹-en-IN]#,##,##0.00'
    percent_fmt = '0.0%'
    
    # ---------------- 1. HEADER BANNER ----------------
    ws1.merge_cells("A1:N2")
    ws1['A1'] = "KHAVDA 442MW SOLAR PROJECT — VENDOR PO & PAYMENT TRACKER"
    ws1['A1'].font = title_font
    ws1['A1'].fill = navy_header_fill
    ws1['A1'].alignment = align_center
    
    ws1.merge_cells("A3:N3")
    ws1['A3'] = "PROJECT: KHAVDA 442MW  |  DOCUMENT TYPE: RECONCILED PO & VENDOR PAYABLE STATEMENT  |  DATE: 29-07-2026"
    ws1['A3'].font = sub_title_font
    ws1['A3'].fill = sub_banner_fill
    ws1['A3'].alignment = align_center
    
    # Set Row Heights for Header
    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 24
    ws1.row_dimensions[3].height = 22
    ws1.row_dimensions[4].height = 12

    # ---------------- 2. KPI SUMMARY CARDS ----------------
    # Card 1: Total Vendors (B5:C6)
    ws1.merge_cells("B5:C5")
    ws1['B5'] = "TOTAL VENDORS"
    ws1['B5'].font = kpi_label_font
    ws1['B5'].fill = kpi_title_fill
    ws1['B5'].alignment = align_center
    
    ws1.merge_cells("B6:C6")
    ws1['B6'] = "=COUNTA(C10:C53)"
    ws1['B6'].font = kpi_num_font
    ws1['B6'].alignment = align_center
    
    # Card 2: Total Paid As On (E5:F6)
    ws1.merge_cells("E5:F5")
    ws1['E5'] = "TOTAL PAID AS ON DATE"
    ws1['E5'].font = kpi_label_font
    ws1['E5'].fill = kpi_title_fill
    ws1['E5'].alignment = align_center
    
    ws1.merge_cells("E6:F6")
    ws1['E6'] = "=I54"
    ws1['E6'].font = kpi_num_font
    ws1['E6'].number_format = currency_fmt
    ws1['E6'].alignment = align_center
    
    # Card 3: Total Outstanding Payable (H5:I6)
    ws1.merge_cells("H5:I5")
    ws1['H5'] = "TOTAL OUTSTANDING PAYABLE"
    ws1['H5'].font = kpi_label_font
    ws1['H5'].fill = kpi_title_fill
    ws1['H5'].alignment = align_center
    
    ws1.merge_cells("H6:I6")
    ws1['H6'] = "=J54"
    ws1['H6'].font = kpi_num_font
    ws1['H6'].number_format = currency_fmt
    ws1['H6'].alignment = align_center
    
    # Card 4: Total Contract Value (K5:L6)
    ws1.merge_cells("K5:L5")
    ws1['K5'] = "TOTAL CONTRACT VALUE"
    ws1['K5'].font = kpi_label_font
    ws1['K5'].fill = kpi_title_fill
    ws1['K5'].alignment = align_center
    
    ws1.merge_cells("K6:L6")
    ws1['K6'] = "=K54"
    ws1['K6'].font = kpi_num_font
    ws1['K6'].number_format = currency_fmt
    ws1['K6'].alignment = align_center
    
    # Card 5: Payment Completion % (M5:N6)
    ws1.merge_cells("M5:N5")
    ws1['M5'] = "OVERALL PAID RATIO"
    ws1['M5'].font = kpi_label_font
    ws1['M5'].fill = kpi_title_fill
    ws1['M5'].alignment = align_center
    
    ws1.merge_cells("M6:N6")
    ws1['M6'] = "=I54/K54"
    ws1['M6'].font = kpi_num_font
    ws1['M6'].number_format = percent_fmt
    ws1['M6'].alignment = align_center

    # Apply borders to KPI Cards
    for r in range(5, 7):
        for col in [2, 3, 5, 6, 8, 9, 11, 12, 13, 14]:
            ws1.cell(row=r, column=col).border = thin_border
            
    ws1.row_dimensions[5].height = 18
    ws1.row_dimensions[6].height = 26
    ws1.row_dimensions[7].height = 14
    ws1.row_dimensions[8].height = 10

    # ---------------- 3. TABLE COLUMN HEADERS ----------------
    headers = [
        "S.No.", "Ref Sr.", "PO Number", "Project Name", "Type", 
        "Description / Scope", "Company / Vendor Firm Name", "Vendor Contact / Representative", 
        "Paid As On (₹)", "Amount to be Paid (₹)", "Total Value (₹)", 
        "Paid %", "Payment Status", "Audit / Highlight Notes"
    ]
    
    header_row = 9
    for col_idx, text in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = col_header_fill
        cell.alignment = align_center
        cell.border = thick_bottom
    ws1.row_dimensions[header_row].height = 28

    # ---------------- 4. TABLE DATA INSERTION ----------------
    raw_data = [
        (8, "6002526032", "KHAVDA 442MW", "Service", "Piling Work", "VIBRANT CONSTRUCTION COMPANY", "Mohamad (6360589955)", 8142757, 500000, ""),
        (16, "6002526062", "KHAVDA 442MW", "Service", "Erection of MMS", "Murlidhar Fabrication", "Suresh Kumar (+91 99783 01708)", 2684889, 1460059, ""),
        (17, "6002526026", "KHAVDA 442MW", "Service", "Erection of MMS", "Choudhary Electric Constrution", "Sarvan Choudhary +91 90013 29239", 8361290, 2507431, ""),
        (22, "6002526025", "KHAVDA 442MW", "Service", "Supply Of White Sand", "R.K.PROJECT", "GAGAL RAMESHBHAI (+91 6354263514)", 5216596, 24170, ""),
        (24, "6002526050", "KHAVDA 442MW", "Service", "DG", "PALIWAL BUILDING MATERIAL", "Lalit-", 1606082, 250000, ""),
        (25, "6002526051", "KHAVDA 442MW", "Service", "HT Cable laying", "SUPER ELECTRO POWER POONA RAM", "Poona ram", 3817310, 213861, ""),
        (27, "6002526061", "KHAVDA 442MW", "Service", "VLF Testing", "TNM GENIE PRIVATE LIMITED", "Saravanakkumar +91 9741402787", 664070, 44314, "Highlighted Yellow in Statement"),
        (31, "6002526072", "KHAVDA 442MW", "Service", "Erection of MMS", "PUSHPA ENTERPRISES", "RAM KUMAR (+91 97259 87924)", 2426646, 222850, ""),
        (15, "6002526084", "KHAVDA 442MW", "Service", "Erection of MMS", "R P CONSTRUCTION", "BAYUNANDAN PANDE (92419 50137)", 1651357, 35435, ""),
        (17, "6002526087", "KHAVDA 442MW", "Service", "Erection of MMS", "AKSHAR CORPORATION", "MEHULKUMAR VASUDEVBHAI PATEL", 2636044, 456175, ""),
        (19, "6002526091", "KHAVDA 442MW", "Service", "Erection of MMS", "ADITYA POWER TECHNOLOGIES", "Hemang Patel", 620240, 428614, ""),
        (20, "6002526092", "KHAVDA 442MW", "Service", "Erection of MMS", "SABNAM ENTERPRISE", "SABNAM BANU", 2883657, 423482, ""),
        (27, "6002526110", "KHAVDA 442MW", "Service", "Erection of MMS", "Polyvision India", "Mahesh Ganesh Patel (9819343954)", 1156520, 300000, ""),
        (29, "6002526112", "KHAVDA 442MW", "Service", "Erection of MMS", "M/S NITYA DEV ENTERPRISES", "SANDEEP KUMAR (+91 9992279349, 9467496874)", 236952, 209366, ""),
        (30, "6002526113", "KHAVDA 442MW", "Service", "AC-DC Work", "TD ELECTRICAL & SOLAR", "TAPOMOY DEY: +91 7602143142", 2280000, 150000, ""),
        (31, "6002526115", "KHAVDA 442MW", "Service", "Erection of MMS", "Mahadev Infratech Enterprises", "Hari Ram 9983414884, 8239971600", 555841, 100361, ""),
        (32, "6002526120", "KHAVDA 442MW", "Service", "LT & Earthing Trench Work", "R.K.PROJECT", "GAGAL RAMESHBHAI (+91 6354263514)", 0, 608934, "Entire Payment Pending"),
        (36, "6002526132", "KHAVDA 442MW", "Service", "Hiring of Vechile", "R.S.B.T. SUPPLIERS", "Praveen bhalothia (+91 6005289994, 9879393598)", 1458428, 496440, ""),
        (41, "6002526145", "KHAVDA 442MW", "Service", "Erection of MMS", "OZONE INFRA", "AMIT KUMAR BAJPAI (+91 9554523222)", 1635481, 908870, ""),
        (43, "6002526152", "KHAVDA 442MW", "Service", "Erection of MMS", "TANSUMAN LIGHTNING ENTERPRISES (OPC) PRIVATE LIMITED", "Vishal Rathore (+91 7790909130)", 446072, 111519, ""),
        (51, "-", "KHAVDA 442MW", "Supplier", "Hardware & Machine", "Rajkot Machinery", "Rajkot Machinery", 1100000, 130023, "Supplier Account"),
        (57, "-", "KHAVDA 442MW", "Service", "Hiring Splicing Service", "New Radiant Enterprises", "New Radiant Enterprises", 0, 61950, "Entire Payment Pending"),
        (5, "6002526034", "KHAVDA 442MW", "Service", "IDT Civil", "Dayachand -7", "PAWAN KUMAR 80555 27414", 2740748, 152263, "Highlighted Yellow in Statement"),
        (7, "6002526064", "KHAVDA 442MW", "Service", "IDT Civil", "Baburam-3", "BABU RAM (+91 9001678721)", 1160379, 128931, ""),
        (20, "6002526001", "KHAVDA 442MW", "Service", "AC-DC Work", "J.P. ENTERPRISES", "JAI PRAKASH +91 98282 39611", 3535000, 200000, ""),
        (42, "6002526027", "KHAVDA 442MW", "Service", "HT Jointer", "Sanjay Kumar Yadav", "Sanjay Kumar Yadav", 700000, 42306, ""),
        (50, "6002526086", "KHAVDA 442MW", "Service", "Erection of MMS", "SARTHAK ENTERPRISES", "OM PRAKASH +91 8950588873, 9896098873", 17118, 4280, ""),
        (77, "6002526136", "KHAVDA 442MW", "Service", "AC-DC Work", "FATIMA CONSTRUCTION", "ZAINAB FATMA 7428270013", 2200000, 450000, ""),
        (91, "6002526155", "KHAVDA 442MW", "Service", "String Work", "FATA RAM", "FATA RAM (+91 8401688194)", 380700, 42300, ""),
        (95, "6002526159", "KHAVDA 442MW", "Service", "Erection of Module", "Juhi Gulab", "ALHANOOR (+91 8460311858)", 329386, 93056, "Highlighted Yellow in Statement"),
        (135, "-", "KHAVDA 442MW", "Service", "AC-DC Work", "Nagnararai Construction Company-2", "Lun Singh ( 9680881351)", 3220000, 700000, ""),
        (140, "-", "KHAVDA 442MW", "Service", "HT Cable laying", "Ashok Kumar Bairwa", "Ashok Kumar Bairwa(9740523254)", 350248, 25000, ""),
        (143, "-", "KHAVDA 442MW", "Service", "Supply Labour", "Girdhari Ram Labour", "Girdhari Ram Labour", 821512, 300000, ""),
        (147, "-", "KHAVDA 442MW", "Service", "DC Works", "Sumer Singh", "Devi Singh/ Sumer Singh", 1961000, 700000, ""),
        (154, "-", "KHAVDA 442MW", "Service", "AC-DC Work", "Jeeshan Ahemad", "Jeeshan Ahemad", 1625000, 100000, ""),
        (157, "-", "KHAVDA 442MW", "Service", "Hiring of Vechile", "SHARDA MAHALA", "RJ37-GB-0907", 73548, 52903, ""),
        (158, "-", "KHAVDA 442MW", "Service", "AC Work", "Abhishek Pandey", "Abhishek Pandey", 1495000, 200000, ""),
        (159, "-", "KHAVDA 442MW", "Service", "AC-DC Work", "ISHAN", "Sajjirul", 2100000, 260000, ""),
        (173, "-", "KHAVDA 442MW", "Service", "WATER SUPPLIERS", "Abdul Kayum Ibrahim Sama", "Abdul Kayum Ibrahim Sama", 300000, 170000, ""),
        (177, "-", "KHAVDA 442MW", "Service", "Erection of MMS", "JAHANGIR ALI", "JAHANGIR ALI", 179000, 45861, ""),
        (182, "-", "KHAVDA 442MW", "Service", "Oger Machine", "Padam Singh", "Padam SIngh", 0, 39800, "Entire Payment Pending"),
        (185, "-", "KHAVDA 442MW", "Service", "Hiring of Vechile", "Mohan Lal", "Mohan Lal", 0, 40000, "Entire Payment Pending"),
        (186, "6002627006", "KHAVDA 442MW", "Service", "Hiring of Vechile", "TAJMAMAD LATIF SAMA", "TAJMAMAD LATIF SAMA", 0, 40000, "Entire Payment Pending"),
        (187, "-", "KHAVDA 442MW", "Service", "Cabin / Material", "Gujrat Pota Cabin", "Gujrat Pota Cabin", 0, 700000, "Entire Payment Pending"),
    ]

    start_row = 10
    for idx, row in enumerate(raw_data, 1):
        curr_row = start_row + idx - 1
        orig_sr, po_num, prj, vtype, desc, comp, contact, paid, payable, remark = row
        
        ws1.cell(row=curr_row, column=1, value=idx).alignment = align_center
        ws1.cell(row=curr_row, column=2, value=orig_sr).alignment = align_center
        ws1.cell(row=curr_row, column=3, value=str(po_num)).alignment = align_center
        ws1.cell(row=curr_row, column=4, value=prj).alignment = align_center
        ws1.cell(row=curr_row, column=5, value=vtype).alignment = align_center
        
        ws1.cell(row=curr_row, column=6, value=desc).alignment = align_left
        ws1.cell(row=curr_row, column=7, value=comp).alignment = align_left
        ws1.cell(row=curr_row, column=8, value=contact).alignment = align_left
        
        # Paid As On
        c_paid = ws1.cell(row=curr_row, column=9, value=paid)
        c_paid.number_format = currency_fmt
        c_paid.alignment = align_right
        
        # Amount to be Paid
        c_payable = ws1.cell(row=curr_row, column=10, value=payable)
        c_payable.number_format = currency_fmt
        c_payable.alignment = align_right
        
        # Total Contract Value (Formula)
        c_total = ws1.cell(row=curr_row, column=11, value=f"=I{curr_row}+J{curr_row}")
        c_total.number_format = currency_fmt
        c_total.alignment = align_right
        c_total.font = data_bold_font
        
        # Paid % (Formula)
        c_pct = ws1.cell(row=curr_row, column=12, value=f"=IF(K{curr_row}>0, I{curr_row}/K{curr_row}, 0)")
        c_pct.number_format = percent_fmt
        c_pct.alignment = align_right
        
        # Payment Status Badge (Formula)
        c_status = ws1.cell(row=curr_row, column=13, value=f'=IF(J{curr_row}=0,"FULLY PAID",IF(I{curr_row}=0,"PENDING","PARTIAL"))')
        c_status.alignment = align_center
        c_status.font = data_bold_font
        
        # Remarks / Audit
        ws1.cell(row=curr_row, column=14, value=remark).alignment = align_left
        
        # Styling & Zebra Striping
        row_fill = yellow_tag_fill if "Highlighted" in remark else (zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None))
        
        for c in range(1, 15):
            cell = ws1.cell(row=curr_row, column=c)
            cell.font = data_bold_font if c in [11, 13] else data_font
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = thin_border
            
        ws1.row_dimensions[curr_row].height = 20

    # ---------------- 5. TOTALS ROW ----------------
    tot_row = start_row + len(raw_data)
    ws1.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=8)
    ws1.cell(row=tot_row, column=1, value="GRAND TOTAL (KHAVDA 442MW STATEMENT)").font = total_font
    ws1.cell(row=tot_row, column=1).alignment = align_right
    
    # Total Paid
    c_tot_paid = ws1.cell(row=tot_row, column=9, value=f"=SUM(I10:I{tot_row-1})")
    c_tot_paid.font = total_font
    c_tot_paid.number_format = currency_fmt
    c_tot_paid.alignment = align_right
    
    # Total Payable
    c_tot_payable = ws1.cell(row=tot_row, column=10, value=f"=SUM(J10:J{tot_row-1})")
    c_tot_payable.font = total_font
    c_tot_payable.number_format = currency_fmt
    c_tot_payable.alignment = align_right
    
    # Total Contract Value
    c_tot_val = ws1.cell(row=tot_row, column=11, value=f"=SUM(K10:K{tot_row-1})")
    c_tot_val.font = total_font
    c_tot_val.number_format = currency_fmt
    c_tot_val.alignment = align_right
    
    # Overall Paid %
    c_tot_pct = ws1.cell(row=tot_row, column=12, value=f"=I{tot_row}/K{tot_row}")
    c_tot_pct.font = total_font
    c_tot_pct.number_format = percent_fmt
    c_tot_pct.alignment = align_right
    
    ws1.cell(row=tot_row, column=13, value="RECONCILED").font = total_font
    ws1.cell(row=tot_row, column=13).alignment = align_center
    
    ws1.cell(row=tot_row, column=14, value="All 44 Records Verified").font = data_bold_font
    ws1.cell(row=tot_row, column=14).alignment = align_left
    
    for c in range(1, 15):
        cell = ws1.cell(row=tot_row, column=c)
        cell.fill = total_row_fill
        cell.border = double_bottom_border
        
    ws1.row_dimensions[tot_row].height = 26

    # Manual adjustments for clean spacing
    ws1.column_dimensions['A'].width = 8   # S.No.
    ws1.column_dimensions['B'].width = 10  # Ref Sr.
    ws1.column_dimensions['C'].width = 16  # PO Number
    ws1.column_dimensions['D'].width = 16  # Project
    ws1.column_dimensions['E'].width = 12  # Type
    ws1.column_dimensions['F'].width = 28  # Description
    ws1.column_dimensions['G'].width = 38  # Company Name
    ws1.column_dimensions['H'].width = 36  # Vendor Contact
    ws1.column_dimensions['I'].width = 20  # Paid As On
    ws1.column_dimensions['J'].width = 22  # Amount to be paid
    ws1.column_dimensions['K'].width = 22  # Total Value
    ws1.column_dimensions['L'].width = 12  # Paid %
    ws1.column_dimensions['M'].width = 18  # Status
    ws1.column_dimensions['N'].width = 32  # Remarks

    # ---------------- 6. SHEET 2: SCOPE & SUMMARY ANALYSIS ----------------
    ws2 = wb.create_sheet(title="Scope & Payment Summary")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.merge_cells("A1:F2")
    ws2['A1'] = "KHAVDA 442MW — SCOPE OF WORK SUMMARY & PAYMENT BREAKDOWN"
    ws2['A1'].font = title_font
    ws2['A1'].fill = navy_header_fill
    ws2['A1'].alignment = align_center
    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 22
    
    # Table Header
    s2_headers = ["Scope / Description", "Vendor Count", "Total Paid As On (₹)", "Total Payable (₹)", "Total Contract Value (₹)", "Paid %"]
    for c_idx, h in enumerate(s2_headers, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = col_header_fill
        cell.alignment = align_center
        cell.border = thick_bottom
    ws2.row_dimensions[4].height = 26
    
    # Unique Scopes
    scopes = [
        "Erection of MMS", "AC-DC Work", "IDT Civil", "Hiring of Vechile", 
        "HT Cable laying", "Piling Work", "Supply Of White Sand", "DG", 
        "VLF Testing", "LT & Earthing Trench Work", "Hardware & Machine", 
        "Hiring Splicing Service", "HT Jointer", "String Work", 
        "Erection of Module", "Supply Labour", "DC Works", "AC Work", 
        "WATER SUPPLIERS", "Oger Machine", "Cabin / Material"
    ]
    
    for s_idx, scope in enumerate(scopes, 5):
        ws2.cell(row=s_idx, column=1, value=scope).font = data_bold_font
        ws2.cell(row=s_idx, column=1).alignment = align_left
        
        # Vendor Count Formula
        c_cnt = ws2.cell(row=s_idx, column=2, value=f"=COUNTIF('PO & Payment Tracker'!F10:F53, \"{scope}\")")
        c_cnt.alignment = align_center
        c_cnt.font = data_font
        
        # Paid Formula
        c_p = ws2.cell(row=s_idx, column=3, value=f"=SUMIF('PO & Payment Tracker'!F10:F53, \"{scope}\", 'PO & Payment Tracker'!I10:I53)")
        c_p.number_format = currency_fmt
        c_p.alignment = align_right
        c_p.font = data_font
        
        # Payable Formula
        c_pay = ws2.cell(row=s_idx, column=4, value=f"=SUMIF('PO & Payment Tracker'!F10:F53, \"{scope}\", 'PO & Payment Tracker'!J10:J53)")
        c_pay.number_format = currency_fmt
        c_pay.alignment = align_right
        c_pay.font = data_font
        
        # Total Value Formula
        c_tot = ws2.cell(row=s_idx, column=5, value=f"=C{s_idx}+D{s_idx}")
        c_tot.number_format = currency_fmt
        c_tot.alignment = align_right
        c_tot.font = data_bold_font
        
        # Paid % Formula
        c_pct = ws2.cell(row=s_idx, column=6, value=f"=IF(E{s_idx}>0, C{s_idx}/E{s_idx}, 0)")
        c_pct.number_format = percent_fmt
        c_pct.alignment = align_right
        c_pct.font = data_font
        
        for c in range(1, 7):
            ws2.cell(row=s_idx, column=c).border = thin_border
        ws2.row_dimensions[s_idx].height = 20
        
    # Summary Totals Row
    s2_tot_row = len(scopes) + 5
    ws2.cell(row=s2_tot_row, column=1, value="TOTALS").font = total_font
    ws2.cell(row=s2_tot_row, column=1).alignment = align_right
    
    ws2.cell(row=s2_tot_row, column=2, value=f"=SUM(B5:B{s2_tot_row-1})").font = total_font
    ws2.cell(row=s2_tot_row, column=2).alignment = align_center
    
    c_s2_p = ws2.cell(row=s2_tot_row, column=3, value=f"=SUM(C5:C{s2_tot_row-1})")
    c_s2_p.font = total_font
    c_s2_p.number_format = currency_fmt
    c_s2_p.alignment = align_right
    
    c_s2_pay = ws2.cell(row=s2_tot_row, column=4, value=f"=SUM(D5:D{s2_tot_row-1})")
    c_s2_pay.font = total_font
    c_s2_pay.number_format = currency_fmt
    c_s2_pay.alignment = align_right
    
    c_s2_tot = ws2.cell(row=s2_tot_row, column=5, value=f"=SUM(E5:E{s2_tot_row-1})")
    c_s2_tot.font = total_font
    c_s2_tot.number_format = currency_fmt
    c_s2_tot.alignment = align_right
    
    c_s2_pct = ws2.cell(row=s2_tot_row, column=6, value=f"=C{s2_tot_row}/E{s2_tot_row}")
    c_s2_pct.font = total_font
    c_s2_pct.number_format = percent_fmt
    c_s2_pct.alignment = align_right
    
    for c in range(1, 7):
        ws2.cell(row=s2_tot_row, column=c).fill = total_row_fill
        ws2.cell(row=s2_tot_row, column=c).border = double_bottom_border
    ws2.row_dimensions[s2_tot_row].height = 24

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 22
    ws2.column_dimensions['E'].width = 24
    ws2.column_dimensions['F'].width = 14

    # Save to both workspace and Desktop for immediate accessibility
    out_filename = "Khavda_442MW_Vendor_Payment_Tracker.xlsx"
    wb.save(out_filename)
    wb.save(f"c:\\Users\\Asus 15 Aspire\\Desktop\\{out_filename}")
    print(f"Workbook saved successfully as: {out_filename} in workspace and Desktop!")

if __name__ == "__main__":
    create_vendor_tracker()

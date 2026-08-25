import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def create_jain_motors_invoice():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Challan 120"

    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Fonts
    FONT_NAME = "Segoe UI"
    font_store_name = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
    font_store_sub = Font(name=FONT_NAME, size=10, italic=True, color="F8F9FA")
    font_section_lbl = Font(name=FONT_NAME, size=10, bold=True, color="495057")
    font_section_val = Font(name=FONT_NAME, size=10, color="212529")
    font_header = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_NAME, size=10, color="212529")
    font_body_bold = Font(name=FONT_NAME, size=10, bold=True, color="212529")
    font_footer = Font(name=FONT_NAME, size=9, italic=True, color="6C757D")
    font_sig_title = Font(name=FONT_NAME, size=9, bold=True, color="495057")

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Borders
    border_color_gray = "D9D9D9"
    border_color_dark = "212529"
    
    side_thin = Side(border_style="thin", color=border_color_gray)
    side_thick_bottom = Side(border_style="medium", color=border_color_dark)
    side_double = Side(border_style="double", color=border_color_dark)
    side_top_thin = Side(border_style="thin", color=border_color_dark)
    
    border_cell = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    border_header = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thick_bottom)
    border_summary = Border(top=side_top_thin, bottom=side_double)

    # Color Fills
    fill_store_header = PatternFill(start_color="212529", end_color="212529", fill_type="solid") # Dark Charcoal
    fill_table_header = PatternFill(start_color="495057", end_color="495057", fill_type="solid") # Slate Gray
    fill_zebra_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_zebra_odd = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    fill_summary = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")       # Light Gray for totals

    # Helper function to style ranges
    def style_range(ws, cell_range, font=None, alignment=None, fill=None, border=None):
        for row in ws[cell_range]:
            for cell in row:
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border

    # Set column widths
    column_widths = {
        "A": 3,   # Left margin
        "B": 8,   # S.No.
        "C": 35,  # Particulars
        "D": 10,  # Qty
        "E": 14,  # Rate
        "F": 18,  # Amount
        "G": 3    # Right margin
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row Heights
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 12
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 15

    # 1. Main Header Title
    ws.merge_cells("B2:F2")
    ws["B2"] = "JAIN MOTORS"
    style_range(ws, "B2:F2", font=font_store_name, alignment=align_center, fill=fill_store_header)

    ws.merge_cells("B3:F3")
    ws["B3"] = "CHALLAN BOOK  |  Mob. : 9358993882"
    style_range(ws, "B3:F3", font=font_store_sub, alignment=align_center, fill=fill_store_header)

    # 2. Metadata details (Challan No, Date, Customer)
    # Row 5
    ws["B5"] = "Challan No:"
    ws["B5"].font = font_section_lbl
    ws["B5"].alignment = align_left
    
    ws["C5"] = 120
    ws["C5"].font = font_section_val
    ws["C5"].alignment = align_left
    
    ws["E5"] = "Date:"
    ws["E5"].font = font_section_lbl
    ws["E5"].alignment = align_right
    
    ws["F5"] = "2026-06-13" # Preserved date or current date
    ws["F5"].font = font_section_val
    ws["F5"].alignment = align_left
    
    # Row 6
    ws["B6"] = "Customer:"
    ws["B6"].font = font_section_lbl
    ws["B6"].alignment = align_left
    
    ws["C6"] = "MSC"
    ws["C6"].font = font_section_val
    ws["C6"].alignment = align_left

    # Draw border around the metadata block
    style_range(ws, "B5:F6", border=Border(bottom=side_thin))

    # 3. Main Table Headers (Row 8)
    headers = [
        ("S.No.", align_center),
        ("Particulars", align_left),
        ("Qty", align_center),
        ("Rate", align_right),
        ("Amount", align_right)
    ]
    
    ws.row_dimensions[8].height = 26
    for idx, (text, align) in enumerate(headers, start=2):
        cell = ws.cell(row=8, column=idx, value=text)
        cell.font = font_header
        cell.alignment = align
        cell.fill = fill_table_header
        cell.border = border_header

    # 4. Transcribed Data (Rows 9 to 18)
    items_data = [
        ["FILTER PANA", 1, 350.00],
        ["BOSAICAL 12 TON", 8, 1200.00],
        ["BOSAICAL 17 TON", 4, 1700.00],
        ["D-CIACL 1 TON", 4, 100.00],
        ["3KG HAMMER", 1, 600.00],
        ["5KG HAMMER", 1, 1000.00],
        ["FAUDA", 2, 250.00],
        ["12x12x36", 20, 3000.00],
        ["4x4x36", 40, 800.00],
        ["TOMMi", 2, 750.00]
    ]

    for i, data in enumerate(items_data):
        row_idx = 9 + i
        ws.row_dimensions[row_idx].height = 22
        fill_zebra = fill_zebra_odd if row_idx % 2 == 1 else fill_zebra_even
        
        # S.No.
        ws.cell(row=row_idx, column=2, value=i+1).alignment = align_center
        # Particulars
        ws.cell(row=row_idx, column=3, value=data[0]).alignment = align_left
        # Qty
        ws.cell(row=row_idx, column=4, value=data[1]).alignment = align_center
        ws.cell(row=row_idx, column=4).number_format = "#,##0"
        # Rate
        ws.cell(row=row_idx, column=5, value=data[2]).alignment = align_right
        ws.cell(row=row_idx, column=5).number_format = "₹#,##0.00"
        # Amount (Formula)
        amt_cell = ws.cell(row=row_idx, column=6, value=f"=D{row_idx}*E{row_idx}")
        amt_cell.alignment = align_right
        amt_cell.number_format = "₹#,##0.00"
        
        # Apply borders, font and fill
        for col_idx in range(2, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_body
            cell.border = border_cell
            cell.fill = fill_zebra

    # 5. Totals Block (Rows 19 to 21)
    ws.row_dimensions[19].height = 22
    ws.row_dimensions[20].height = 22
    ws.row_dimensions[21].height = 24

    # Subtotal
    ws.merge_cells("B19:E19")
    ws["B19"] = "Subtotal"
    style_range(ws, "B19:E19", font=font_body_bold, alignment=align_right, fill=fill_summary, border=border_cell)
    
    subtotal_cell = ws.cell(row=19, column=6, value="=SUM(F9:F18)")
    subtotal_cell.font = font_body_bold
    subtotal_cell.alignment = align_right
    subtotal_cell.number_format = "₹#,##0.00"
    subtotal_cell.fill = fill_summary
    subtotal_cell.border = border_cell

    # GST (18%)
    ws.merge_cells("B20:E20")
    ws["B20"] = "GST (18%)"
    style_range(ws, "B20:E20", font=font_body_bold, alignment=align_right, fill=fill_summary, border=border_cell)
    
    gst_cell = ws.cell(row=20, column=6, value="=F19*0.18")
    gst_cell.font = font_body_bold
    gst_cell.alignment = align_right
    gst_cell.number_format = "₹#,##0.00"
    gst_cell.fill = fill_summary
    gst_cell.border = border_cell

    # Grand Total
    ws.merge_cells("B21:E21")
    ws["B21"] = "Grand Total"
    style_range(ws, "B21:E21", font=font_body_bold, alignment=align_right, fill=fill_summary, border=border_summary)
    
    total_cell = ws.cell(row=21, column=6, value="=F19+F20")
    total_cell.font = Font(name=FONT_NAME, size=11, bold=True, color="212529")
    total_cell.alignment = align_right
    total_cell.number_format = "₹#,##0.00"
    total_cell.fill = fill_summary
    total_cell.border = border_summary

    # Spacer row
    ws.row_dimensions[22].height = 15

    # 6. Brand Footers Block
    ws.merge_cells("B23:F23")
    ws["B23"] = "Authorized Dealer: FERRETERRO | BOSCH | TAPARIA | INGCO"
    style_range(ws, "B23:F23", font=font_footer, alignment=align_center)

    ws.row_dimensions[24].height = 10
    ws.row_dimensions[25].height = 20
    ws.row_dimensions[26].height = 15
    ws.row_dimensions[27].height = 15
    ws.row_dimensions[28].height = 15

    # Signature Block (Bottom Right)
    ws.merge_cells("E25:F25")
    ws["E25"] = "For JAIN MOTORS"
    style_range(ws, "E25:F25", font=font_sig_title, alignment=align_center)

    ws.merge_cells("E27:F27")
    ws["E27"] = "_________________________"
    style_range(ws, "E27:F27", font=font_body, alignment=align_center)

    ws.merge_cells("E28:F28")
    ws["E28"] = "Authorized Signature"
    style_range(ws, "E28:F28", font=font_footer, alignment=align_center)

    # Save Excel Workbook
    output_filename = "Jain_Motors_Challan_120.xlsx"
    wb.save(output_filename)
    print(f"Jain Motors Challan created successfully: {output_filename}")

if __name__ == "__main__":
    create_jain_motors_invoice()

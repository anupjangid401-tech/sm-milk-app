import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def create_tshirt_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T-Shirt Procurement"

    # Set gridlines visible (vital for structured data layout)
    ws.views.sheetView[0].showGridLines = True

    # ---- Design System Tokens & Fonts ----
    FONT_NAME = "Segoe UI"
    
    # Fonts
    font_title = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name=FONT_NAME, size=10, italic=True, color="E0E6ED")
    font_kpi_label = Font(name=FONT_NAME, size=9, bold=True, color="5C6B73")
    font_kpi_value_dark = Font(name=FONT_NAME, size=15, bold=True, color="1B365D")
    font_kpi_value_green = Font(name=FONT_NAME, size=15, bold=True, color="2E7D32")
    font_kpi_value_red = Font(name=FONT_NAME, size=15, bold=True, color="C62828")
    
    font_header = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    font_body = Font(name=FONT_NAME, size=10)
    font_body_bold = Font(name=FONT_NAME, size=10, bold=True)
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Borders
    border_color_gray = "D9D9D9"
    border_color_dark = "1B365D"
    
    side_thin = Side(border_style="thin", color=border_color_gray)
    side_double = Side(border_style="double", color=border_color_dark)
    side_top_thin = Side(border_style="thin", color=border_color_dark)
    
    border_cell = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    border_summary = Border(top=side_top_thin, bottom=side_double)
    border_kpi = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    
    # Color Fills
    fill_header_main = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Deep Navy Blue
    fill_header_table = PatternFill(start_color="2B4C7E", end_color="2B4C7E", fill_type="solid") # Slate/Indigo Blue
    fill_kpi_bg = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")       # Light Indigo Gray
    fill_summary = PatternFill(start_color="E3EBF3", end_color="E3EBF3", fill_type="solid")      # Slate Light
    
    # Zebra striping
    fill_zebra_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_zebra_odd = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")   # Very light tint
    
    # Status styling templates (Fill and Font color)
    status_styles = {
        "Paid": {
            "fill": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),   # Soft Green
            "font": Font(name=FONT_NAME, size=10, bold=True, color="385723")
        },
        "Partially Paid": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),   # Soft Orange/Yellow
            "font": Font(name=FONT_NAME, size=10, bold=True, color="7F6000")
        },
        "Unpaid": {
            "fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),   # Soft Red
            "font": Font(name=FONT_NAME, size=10, bold=True, color="C00000")
        },
        "Local": {
            "fill": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),   # Soft Blue
            "font": Font(name=FONT_NAME, size=10, color="1A73E8")
        },
        "Outstation": {
            "fill": PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid"),   # Soft Gray
            "font": Font(name=FONT_NAME, size=10, color="5F6368")
        },
        "Delivered": {
            "fill": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),   # Soft Green
            "font": Font(name=FONT_NAME, size=10, color="385723")
        },
        "In Transit": {
            "fill": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),   # Soft Blue
            "font": Font(name=FONT_NAME, size=10, color="1A73E8")
        },
        "Pending": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),   # Soft Yellow
            "font": Font(name=FONT_NAME, size=10, color="7F6000")
        }
    }

    # Helper function to style ranges
    def style_range(ws, cell_range, font=None, alignment=None, fill=None, border=None):
        for row in ws[cell_range]:
            for cell in row:
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border

    # ---- 1. Title Block (Rows 2 & 3) ----
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 18
    
    ws.merge_cells("B2:N2")
    ws["B2"] = "T-SHIRT PROCUREMENTS & VENDOR LEDGER"
    style_range(ws, "B2:N2", font=font_title, alignment=align_center, fill=fill_header_main)
    
    ws.merge_cells("B3:N3")
    ws["B3"] = "Track item descriptions, quantities, pricing, payment settlements, and vendor locations."
    style_range(ws, "B3:N3", font=font_subtitle, alignment=align_center, fill=fill_header_main)

    # Spacer
    ws.row_dimensions[4].height = 15

    # ---- 2. KPI Summary Cards (Rows 5 & 6) ----
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 26

    kpi_configs = [
        {"range_label": "B5:D5", "range_val": "B6:D6", "label": "TOTAL T-SHIRTS ORDERED", "formula": "=SUM(G10:G30)", "font_val": font_kpi_value_dark, "num_fmt": "#,##0"},
        {"range_label": "E5:G5", "range_val": "E6:G6", "label": "TOTAL ORDER VALUE", "formula": "=SUM(I10:I30)", "font_val": font_kpi_value_dark, "num_fmt": "₹#,##0.00"},
        {"range_label": "H5:J5", "range_val": "H6:J6", "label": "TOTAL AMOUNT PAID", "formula": "=SUM(K10:K30)", "font_val": font_kpi_value_green, "num_fmt": "₹#,##0.00"},
        {"range_label": "K5:M5", "range_val": "K6:M6", "label": "PENDING BALANCE", "formula": "=SUM(L10:L30)", "font_val": font_kpi_value_red, "num_fmt": "₹#,##0.00"}
    ]

    for config in kpi_configs:
        # Style label
        ws.merge_cells(config["range_label"])
        start_cell_lbl = config["range_label"].split(":")[0]
        ws[start_cell_lbl] = config["label"]
        style_range(ws, config["range_label"], font=font_kpi_label, alignment=align_center, fill=fill_kpi_bg, border=border_kpi)
        
        # Style value
        ws.merge_cells(config["range_val"])
        start_cell_val = config["range_val"].split(":")[0]
        ws[start_cell_val] = config["formula"]
        ws[start_cell_val].number_format = config["num_fmt"]
        style_range(ws, config["range_val"], font=config["font_val"], alignment=align_center, fill=fill_kpi_bg, border=border_kpi)

    # Spacer
    ws.row_dimensions[7].height = 10
    ws.row_dimensions[8].height = 10

    # ---- 3. Main Data Table Header (Row 9) ----
    headers = [
        ("Sr. No.", align_center),
        ("Order Date", align_center),
        ("Material Description", align_left),
        ("Supplier / Vendor", align_left),
        ("Vendor Type", align_center),
        ("Qty Ordered", align_right),
        ("Unit Price", align_right),
        ("Total Amount", align_right),
        ("Payment Status", align_center),
        ("Amount Paid", align_right),
        ("Balance Pending", align_right),
        ("Delivery Status", align_center),
        ("Remarks / Notes", align_left)
    ]
    
    ws.row_dimensions[9].height = 28
    for col_idx, (text, alignment) in enumerate(headers, start=2): # Start from column B (2)
        cell = ws.cell(row=9, column=col_idx, value=text)
        cell.font = font_header
        cell.alignment = alignment
        cell.fill = fill_header_table
        cell.border = border_cell

    # ---- 4. Data Rows (Rows 10 to 30) ----
    mock_data = [
        ["2026-06-01", "Premium Cotton Round Neck T-Shirt (M)", "Sharma Hosiery & Garments", "Local", 500, 180.00, "Paid", 90000.00, "Delivered", "Bulk purchase for summer event."],
        ["2026-06-03", "Polyester Sports Polo T-Shirt (L)", "Vikas Knitwear", "Outstation", 300, 220.00, "Partially Paid", 36000.00, "Delivered", "Advance of Rs 36,000 paid."],
        ["2026-06-05", "V-Neck Bio-wash Cotton T-Shirt (S)", "Apex Uniforms & Co.", "Local", 1000, 165.00, "Unpaid", 0.00, "Pending", "PO issued. Awaiting delivery."],
        ["2026-06-08", "Oversized Graphic T-Shirt (XL)", "Classic Prints Vendor", "Local", 200, 250.00, "Paid", 50000.00, "Delivered", "Full payment made via UPI."],
        ["2026-06-10", "Full Sleeve Winter Polo T-Shirt (M)", "Elite Apparels", "Outstation", 400, 290.00, "Partially Paid", 60000.00, "In Transit", "Balance post delivery confirmation."]
    ]

    total_rows = 21  # 5 mock rows + 16 blank rows (Rows 10 to 30)
    for i in range(total_rows):
        row_idx = 10 + i
        ws.row_dimensions[row_idx].height = 20
        fill_zebra = fill_zebra_odd if row_idx % 2 == 1 else fill_zebra_even
        
        # Col B: Sr No
        ws.cell(row=row_idx, column=2, value=i+1).alignment = align_center
        
        if i < len(mock_data):
            # Fill Mock Data
            data = mock_data[i]
            ws.cell(row=row_idx, column=3, value=data[0]).alignment = align_center # Date
            ws.cell(row=row_idx, column=4, value=data[1]).alignment = align_left   # Desc
            ws.cell(row=row_idx, column=5, value=data[2]).alignment = align_left   # Vendor
            
            # Vendor Type + styling
            v_cell = ws.cell(row=row_idx, column=6, value=data[3])
            v_cell.alignment = align_center
            if data[3] in status_styles:
                v_cell.fill = status_styles[data[3]]["fill"]
                v_cell.font = status_styles[data[3]]["font"]
                
            ws.cell(row=row_idx, column=7, value=data[4]).alignment = align_right  # Qty
            ws.cell(row=row_idx, column=7).number_format = "#,##0"
            
            ws.cell(row=row_idx, column=8, value=data[5]).alignment = align_right  # Unit Price
            ws.cell(row=row_idx, column=8).number_format = "₹#,##0.00"
            
            # Total Amount (Formula)
            total_amt_cell = ws.cell(row=row_idx, column=9, value=f"=G{row_idx}*H{row_idx}")
            total_amt_cell.alignment = align_right
            total_amt_cell.number_format = "₹#,##0.00"
            
            # Payment Status + styling
            p_cell = ws.cell(row=row_idx, column=10, value=data[6])
            p_cell.alignment = align_center
            if data[6] in status_styles:
                p_cell.fill = status_styles[data[6]]["fill"]
                p_cell.font = status_styles[data[6]]["font"]
                
            ws.cell(row=row_idx, column=11, value=data[7]).alignment = align_right # Amount Paid
            ws.cell(row=row_idx, column=11).number_format = "₹#,##0.00"
            
            # Balance Pending (Formula)
            bal_cell = ws.cell(row=row_idx, column=12, value=f"=I{row_idx}-K{row_idx}")
            bal_cell.alignment = align_right
            bal_cell.number_format = "₹#,##0.00"
            
            # Delivery Status + styling
            d_cell = ws.cell(row=row_idx, column=13, value=data[8])
            d_cell.alignment = align_center
            if data[8] in status_styles:
                d_cell.fill = status_styles[data[8]]["fill"]
                d_cell.font = status_styles[data[8]]["font"]
                
            ws.cell(row=row_idx, column=14, value=data[9]).alignment = align_left  # Remarks
            
        else:
            # Empty rows pre-formatted
            # Set formulas for Total and Balance, but hide if Quantity is empty
            ws.cell(row=row_idx, column=3).alignment = align_center # Date
            ws.cell(row=row_idx, column=4).alignment = align_left   # Desc
            ws.cell(row=row_idx, column=5).alignment = align_left   # Vendor
            ws.cell(row=row_idx, column=6).alignment = align_center # Vendor Type
            
            # Qty and Rate inputs are empty, formatted
            ws.cell(row=row_idx, column=7).alignment = align_right
            ws.cell(row=row_idx, column=7).number_format = "#,##0"
            ws.cell(row=row_idx, column=8).alignment = align_right
            ws.cell(row=row_idx, column=8).number_format = "₹#,##0.00"
            
            # Total Formula: =IF(G14="","",G14*H14)
            tot_cell = ws.cell(row=row_idx, column=9, value=f'=IF(G{row_idx}="","",G{row_idx}*H{row_idx})')
            tot_cell.alignment = align_right
            tot_cell.number_format = "₹#,##0.00"
            
            ws.cell(row=row_idx, column=10).alignment = align_center # Payment Status
            
            # Paid Amount input is empty, formatted
            ws.cell(row=row_idx, column=11).alignment = align_right
            ws.cell(row=row_idx, column=11).number_format = "₹#,##0.00"
            
            # Balance Pending Formula: =IF(I14="","",I14-K14)
            bal_cell = ws.cell(row=row_idx, column=12, value=f'=IF(I{row_idx}="","",I{row_idx}-K{row_idx})')
            bal_cell.alignment = align_right
            bal_cell.number_format = "₹#,##0.00"
            
            ws.cell(row=row_idx, column=13).alignment = align_center # Delivery Status
            ws.cell(row=row_idx, column=14).alignment = align_left   # Remarks
            
        # Apply standard styling to all data rows
        for col_idx in range(2, 15):
            c = ws.cell(row=row_idx, column=col_idx)
            # Apply zebra background if not already colored by status values
            if c.fill.fill_type is None:
                c.fill = fill_zebra
            if c.font == Font(): # Apply default body font if not overridden
                c.font = font_body
            c.border = border_cell

    # ---- 5. Totals Row (Row 31) ----
    ws.row_dimensions[31].height = 24
    
    # Merge label cell B31:F31
    ws.merge_cells("B31:F31")
    ws["B31"] = "Total"
    style_range(ws, "B31:F31", font=font_body_bold, alignment=align_right, fill=fill_summary, border=border_summary)
    
    # Total Qty
    qty_tot = ws.cell(row=31, column=7, value="=SUM(G10:G30)")
    qty_tot.font = font_body_bold
    qty_tot.alignment = align_right
    qty_tot.number_format = "#,##0"
    qty_tot.fill = fill_summary
    qty_tot.border = border_summary
    
    # Empty column H31 (Rate)
    rate_tot = ws.cell(row=31, column=8, value="")
    rate_tot.fill = fill_summary
    rate_tot.border = border_summary
    
    # Total Amount
    amt_tot = ws.cell(row=31, column=9, value="=SUM(I10:I30)")
    amt_tot.font = font_body_bold
    amt_tot.alignment = align_right
    amt_tot.number_format = "₹#,##0.00"
    amt_tot.fill = fill_summary
    amt_tot.border = border_summary
    
    # Empty payment status column J31
    p_tot = ws.cell(row=31, column=10, value="")
    p_tot.fill = fill_summary
    p_tot.border = border_summary
    
    # Total Paid
    paid_tot = ws.cell(row=31, column=11, value="=SUM(K10:K30)")
    paid_tot.font = font_body_bold
    paid_tot.alignment = align_right
    paid_tot.number_format = "₹#,##0.00"
    paid_tot.fill = fill_summary
    paid_tot.border = border_summary
    
    # Total Balance Pending
    bal_tot = ws.cell(row=31, column=12, value="=SUM(L10:L30)")
    bal_tot.font = font_body_bold
    bal_tot.alignment = align_right
    bal_tot.number_format = "₹#,##0.00"
    bal_tot.fill = fill_summary
    bal_tot.border = border_summary
    
    # Empty delivery status and remarks columns M31:N31
    for col_idx in [13, 14]:
        c = ws.cell(row=31, column=col_idx, value="")
        c.fill = fill_summary
        c.border = border_summary

    # Set exact widths
    column_widths = {
        "A": 3,
        "B": 8,
        "C": 14,
        "D": 38,
        "E": 26,
        "F": 14,
        "G": 14,
        "H": 15,
        "I": 18,
        "J": 16,
        "K": 18,
        "L": 18,
        "M": 16,
        "N": 30
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save output file
    output_filename = "T_Shirt_Purchase_Tracker.xlsx"
    wb.save(output_filename)
    print(f"Excel tracker successfully created: {output_filename}")

if __name__ == "__main__":
    create_tshirt_tracker()

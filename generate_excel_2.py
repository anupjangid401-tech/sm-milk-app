import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_invoice_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    font_company = Font(name="Calibri", size=16, bold=True)
    font_subtitle = Font(name="Calibri", size=11, italic=False)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_header = Font(name="Calibri", size=11, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # 1. Company Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "JAISAL ENGINEER & COMPANY"
    ws["A1"].font = font_company
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:F2")
    ws["A2"] = "Pansari Bazar, Jaisalmer(raj.)"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_center

    ws.merge_cells("A3:F3")
    ws["A3"] = "345001"
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = align_center

    ws.merge_cells("A4:F4")
    ws["A4"] = "Mobile No.-9414661124"
    ws["A4"].font = font_subtitle
    ws["A4"].alignment = align_center

    ws.merge_cells("A5:F5")
    ws["A5"] = "GST IN. :-08ADIPG0413C1ZA"
    ws["A5"].font = font_bold
    ws["A5"].alignment = align_center

    # Divider row
    ws.row_dimensions[6].height = 15

    # 2. INVOICE Title
    ws.merge_cells("A7:F7")
    ws["A7"] = "INVOICE"
    ws["A7"].font = font_bold
    ws["A7"].alignment = align_center
    ws["A7"].border = thin_border

    # 3. Customer & Invoice Info Box
    # A8:C12 is Customer Info, D8:F12 is Invoice Info
    customer_info = [
        ("Customer Name :- MS CHOUHAN", ""),
        ("INFRAVENTURE PVT. LTD.", ""),
        ("", ""),
        ("Contact No.:", ""),
        ("Mobile:", "")
    ]
    
    invoice_info = [
        ("Invoice No.-6543", "Dated:16.05.2026"),
        ("Our Ref.Ref.No.", "Dated:"),
        ("Custmer Ref.No.", "Dated:"),
        ("Kind Attn:- Mr.", ""),
        ("", "")
    ]

    for idx, (label, val) in enumerate(customer_info):
        row = 8 + idx
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=label).font = font_bold if idx == 0 else font_regular
        ws.cell(row=row, column=1).alignment = align_left

    for idx, (col_d_val, col_e_val) in enumerate(invoice_info):
        row = 8 + idx
        # Column D
        ws.cell(row=row, column=4, value=col_d_val).font = font_regular
        ws.cell(row=row, column=4).alignment = align_left
        # Column E & F merged for dates/values
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=col_e_val).font = font_regular
        ws.cell(row=row, column=5).alignment = align_left

    # Draw border around customer/invoice info box
    for r in range(7, 13):
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            top = thin_side if r == 7 else None
            bottom = thin_side if r == 12 else None
            left = thin_side if c == 1 or c == 4 else None
            right = thin_side if c == 3 or c == 6 else None
            
            cell.border = Border(
                left=left or cell.border.left,
                right=right or cell.border.right,
                top=top or cell.border.top,
                bottom=bottom or cell.border.bottom
            )

    # 4. Table Header
    headers = ["Sr.No.", "Item Code & Description", "Unit", "Qty", "Basic Rate", "Amount"]
    ws.row_dimensions[13].height = 25
    for c_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=c_idx, value=header)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    # 5. Table Data
    data = [
        (1, "16amp male - female", "Nos.", 11, 500.00),
        (2, "32amp Spin male", "Nos.", 2, 350.00),
        (3, "ABB MCB TP 16amp", "Nos.", 1, 1250.00),
        (4, "Ring Spanner 6x7 to 30x32 Model No.1812", "Nos.", 1, 2067.00),
        (5, "Fix Spanner 6x7 to 30x32 Model No.DEP12", "Nos.", 1, 975.00),
        (6, "Distribustion Board", "Nos.", 2, 5980.00),
        (7, "LED Flood Light 200w", "Nos.", 4, 3400.00),
        (8, "Power Cable 2.5sqmm x 3core", "Mtr.", 180, 140.00),
        (9, "Cable Cutter CC32", "Nos.", 1, 4280.00),
        (10, "Ins. Tape", "Nos.", 90, 15.00),
        (11, "Taparia Tester", "Nos.", 21, 61.00),
        (12, "Taparia Plier", "Nos.", 1, 303.00),
        (13, "Mesuring Tape 5mtr", "Nos.", 2, 120.00),
        (14, "Heat Gun Dewalt", "Nos.", 2, 6510.00),
        (15, "Tool Bag", "Nos.", 1, 750.00),
        (16, "Socket 1/2 No.15", "Nos.", 2, 75.00),
        (17, "Socket 1/2 No.16", "Nos.", 2, 75.00),
        (18, "Socket 1/2 No.18", "Nos.", 2, 83.00),
        (19, "Socket 1/2 No.12", "Nos.", 2, 75.00),
        (20, "Socket 1/2 No.13", "Nos.", 2, 75.00),
        (21, "Socket 1/2 No.24", "Nos.", 2, 121.00),
        (22, "Socket 1/2 No.32", "Nos.", 2, 190.00),
        (23, "Imapact Socket 1/2 No.17", "Nos.", 1, 117.00),
        (24, "Imapact Socket 1/2 No.19", "Nos.", 2, 134.00),
        (25, "Imapact Socket Long 1/2 No.17", "Nos.", 1, 218.00),
        (26, "Impact Socket 3/4 No.36", "Nos.", 2, 599.00),
        (27, "Taflon Tape", "Nos.", 1, 20.00),
        (28, "Drill Machine 13mm", "Nos.", 2, 3850.00)
    ]

    for idx, row_data in enumerate(data):
        row_num = 14 + idx
        ws.row_dimensions[row_num].height = 20
        
        # Sr No
        ws.cell(row=row_num, column=1, value=row_data[0]).alignment = align_center
        # Description
        ws.cell(row=row_num, column=2, value=row_data[1]).alignment = align_left
        # Unit
        ws.cell(row=row_num, column=3, value=row_data[2]).alignment = align_center
        # Qty
        ws.cell(row=row_num, column=4, value=row_data[3]).alignment = align_center
        # Basic Rate
        ws.cell(row=row_num, column=5, value=row_data[4]).alignment = align_right
        ws.cell(row=row_num, column=5).number_format = "0.00"
        
        # Amount (using formula =Qty * Basic Rate, i.e., =D{row}*E{row})
        amount_cell = ws.cell(row=row_num, column=6, value=f"=D{row_num}*E{row_num}")
        amount_cell.alignment = align_right
        amount_cell.number_format = "0.00"
        
        # Formatting borders & fonts
        for col_idx in range(1, 7):
            c_cell = ws.cell(row=row_num, column=col_idx)
            c_cell.font = font_regular
            c_cell.border = thin_border

    # 6. Summary Rows
    total_start_row = 14 + len(data) # 42
    
    # Round Off
    ws.merge_cells(start_row=total_start_row, start_column=1, end_row=total_start_row, end_column=5)
    ws.cell(row=total_start_row, column=1, value="Round Off").font = font_bold
    ws.cell(row=total_start_row, column=1).alignment = align_right
    ws.cell(row=total_start_row, column=6, value=0.00).font = font_bold
    ws.cell(row=total_start_row, column=6).alignment = align_right
    ws.cell(row=total_start_row, column=6).number_format = "0.00"
    
    # Borders for Round Off row
    for col_idx in range(1, 7):
        ws.cell(row=total_start_row, column=col_idx).border = thin_border

    # Total Net Amount
    net_amount_row = total_start_row + 1 # 43
    ws.merge_cells(start_row=net_amount_row, start_column=1, end_row=net_amount_row, end_column=5)
    ws.cell(row=net_amount_row, column=1, value="Total Net Amount").font = font_bold
    ws.cell(row=net_amount_row, column=1).alignment = align_right
    
    # Formula for Net Amount = SUM(F14:F41) + F42
    ws.cell(row=net_amount_row, column=6, value=f"=SUM(F14:F{total_start_row-1})+F{total_start_row}").font = font_bold
    ws.cell(row=net_amount_row, column=6).alignment = align_right
    ws.cell(row=net_amount_row, column=6).number_format = "0.00"

    # Borders for Total Net Amount row
    for col_idx in range(1, 7):
        ws.cell(row=net_amount_row, column=col_idx).border = thin_border

    # Column Width Auto-Fit
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 8:
                continue
            val_str = str(cell.value or "")
            if cell.number_format == "0.00" and isinstance(cell.value, (int, float)):
                val_str = f"{cell.value:.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Adjust specifically for columns
    ws.column_dimensions["A"].width = 8   # Sr.No.
    ws.column_dimensions["B"].width = 45  # Description
    ws.column_dimensions["C"].width = 10  # Unit
    ws.column_dimensions["D"].width = 10  # Qty
    ws.column_dimensions["E"].width = 12  # Basic Rate
    ws.column_dimensions["F"].width = 15  # Amount

    output_path = "c:/Users/Asus 15 Aspire/Desktop/another/Invoice_Jaisal_Engineer_6543.xlsx"
    wb.save(output_path)
    print(f"Excel file generated successfully at: {output_path}")

if __name__ == "__main__":
    create_invoice_excel()

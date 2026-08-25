import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def build_tax_sheet(ws):
    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    font_italic = Font(name="Calibri", size=11, italic=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Border setup
    thin_side = Side(border_style="thin", color="A6A6A6")
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    double_bottom_side = Side(border_style="double", color="000000")
    border_double_bottom = Border(left=thin_side, right=thin_side, top=thin_side, bottom=double_bottom_side)

    # Column Widths
    ws.column_dimensions["A"].width = 40  # Description
    ws.column_dimensions["B"].width = 12  # Detail/Rate
    for col_idx in range(3, 27):
        ws.column_dimensions[get_column_letter(col_idx)].width = 10

    # 1. Title block
    ws["A1"] = "All Figures in Rs. Crores"
    ws["A1"].font = font_italic
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 20

    # 2. Header Rows
    # Row 2: column numbers 1 to 24
    ws.cell(row=2, column=1, value="Year").font = font_bold
    ws.cell(row=2, column=1).alignment = align_left
    ws.cell(row=2, column=1).border = border_thin
    
    # Empty B2
    ws.cell(row=2, column=2).border = border_thin

    # Row 3: FY names
    ws.cell(row=3, column=1, value="FY").font = font_bold
    ws.cell(row=3, column=1).alignment = align_left
    ws.cell(row=3, column=1).border = border_thin
    ws.cell(row=3, column=2).border = border_thin

    fys = [
        "2025-26", "2026-27", "2027-28", "2028-29", "2029-30", "2030-31",
        "2031-32", "2032-33", "2033-34", "2034-35", "2035-36", "2036-37",
        "2037-38", "2038-39", "2039-40", "2040-41", "2041-42", "2042-43",
        "2043-44", "2044-45", "2045-46", "2046-47", "2047-48", "2048-49"
    ]

    for idx in range(24):
        col_idx = 3 + idx
        # Year number
        cell_yr = ws.cell(row=2, column=col_idx, value=idx + 1)
        cell_yr.font = font_bold
        cell_yr.alignment = align_center
        cell_yr.border = border_thin
        
        # FY Name
        cell_fy = ws.cell(row=3, column=col_idx, value=fys[idx])
        cell_fy.font = font_bold
        cell_fy.alignment = align_center
        cell_fy.border = border_thin

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # Raw Inputs Data
    revenue_vals = [83.84, 142.69, 142.69, 142.69] + [117.69]*20
    gbi_vals = [0.25, 0.25, 0.25, 0.25] + [0.00]*20
    rec_vals = [0.00]*24
    interest_vals = [26.82, 52.29, 48.72, 45.14, 41.57, 37.99, 34.41, 30.84, 27.26, 23.69, 20.11, 16.54, 12.96, 9.39, 5.80, 2.22] + [0.00]*8
    
    # Calculate O&M Expenses dynamically with 5% escalation starting from 14.00 in Year 3
    om_vals = [0.00, 0.00]
    current_om = 14.00
    for yr in range(3, 25):
        om_vals.append(round(current_om, 2))
        current_om *= 1.05
        
    it_dep_vals = [0.31, 0.87, 0.74, 0.63, 0.54, 0.47, 0.40, 0.34, 0.29, 0.25, 0.22, 0.19, 0.16, 0.14, 0.12, 0.10, 0.09, 0.08, 0.07, 0.06, 0.05, 0.04, 0.04, 0.03]
    tax_holiday_flag = [1]*10 + [0]*14
    
    # MAT values directly from the image
    mat_vals = [6.39, 13.14, 10.95, 11.56, 12.10, 12.69, 13.27, 13.84, 14.40, 14.95, 15.49, 16.03, 16.55, 17.06, 17.56, 18.05, 18.24, 17.95, 17.64, 17.32, 16.98, 16.63, 16.25, 15.87]
    # MAT credit usage directly from the image
    mat_credit_usage_vals = [0.00]*10 + [4.44, 4.58, 4.71, 4.84, 4.96, 5.08, 5.13, 1.58] + [0.00]*6

    # Rows definitions and formulas
    # Row Number in spreadsheet starts at 4
    rows_def = [
        # (Label, Column B detail, values/formulas, format)
        ("Total Revenue", "", revenue_vals, "0.00"),
        ("GBI", "", gbi_vals, "0.00"),
        ("REC", "", rec_vals, "0.00"),
        ("Less: Admissible Expenditure", "", None, ""), # Subheader row
        ("Interest On Loan", "", interest_vals, "0.00"),
        ("O&M Expenses", "", om_vals, "0.00"),
        ("Depreciation as per I-T Act", "WDV", it_dep_vals, "0.00"),
        ("Total Expenditure", "", "=Interest + O&M + Depreciation", "0.00"),
        ("Profit due to sale of electricity", "", "=Total_Revenue + GBI + REC - Total_Expenditure", "0.00"),
        ("Accumulated Profit/Loss", "", "=Previous_Accumulated + Current_Profit_for_Tax", "0.00"),
        ("Profit for Tax", "", "=IF(Year=1, 55.02, Profit_due_to_sale)", "0.00"),
        ("80 IA", "", tax_holiday_flag, "#,##0"),
        ("Corporate Tax", 0.2600, "=IF(80_IA=1, 0, ROUND(Profit_for_Tax * Corporate_Tax_Rate, 2))", "0.00"),
        ("MAT", 0.2096, mat_vals, "0.00"),
        ("MAT Credit Year Wise", "", "=IF(Corporate_Tax > MAT, 0, MAT - Corporate_Tax)", "0.00"),
        ("MAT Credit Usage", "", mat_credit_usage_vals, "0.00"),
        ("Tax Payable Without MAT Credit", "", "=MAX(Corporate_Tax, MAT)", "0.00"),
        ("Tax Payable After MAT Credit", "", "=Tax_Payable_Without_MAT_Credit - MAT_Credit_Usage", "0.00")
    ]

    # Pre-calculate row numbers to allow forward references in formulas
    row_map = {}
    temp_row = 4
    for item in rows_def:
        label, _, _, _ = item
        row_map[label] = temp_row
        temp_row += 1

    current_row = 4
    for item in rows_def:
        label, detail, source, num_format = item
        
        # Col A: Label
        c_label = ws.cell(row=current_row, column=1, value=label)
        c_label.font = font_bold if label in ["Less: Admissible Expenditure", "Total Expenditure", "Profit due to sale of electricity", "Accumulated Profit/Loss", "Tax Payable Without MAT Credit", "Tax Payable After MAT Credit"] else font_regular
        c_label.alignment = align_left
        c_label.border = border_thin
        
        # Col B: Detail / Rate
        c_detail = ws.cell(row=current_row, column=2)
        c_detail.alignment = align_center
        c_detail.border = border_thin
        if isinstance(detail, (int, float)):
            c_detail.value = detail
            c_detail.number_format = "0.00%"
            c_detail.font = font_regular
        elif detail != "":
            c_detail.value = detail
            c_detail.font = font_italic
            
        # Draw cells for Year 1 to 24
        for idx in range(24):
            col_idx = 3 + idx
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = border_thin
            cell.alignment = align_right
            cell.font = font_bold if label in ["Total Expenditure", "Profit due to sale of electricity", "Accumulated Profit/Loss", "Tax Payable Without MAT Credit", "Tax Payable After MAT Credit"] else font_regular
            if num_format:
                cell.number_format = num_format
            
            # Populate cell values or formulas
            if label == "Less: Admissible Expenditure":
                # Merge the row cells? No, just leave them empty as in the image
                pass
            elif label == "Total Expenditure":
                r_int = row_map["Interest On Loan"]
                r_om = row_map["O&M Expenses"]
                r_dep = row_map["Depreciation as per I-T Act"]
                cell.value = f"=SUM({col_letter}{r_int}:{col_letter}{r_dep})"
            elif label == "Profit due to sale of electricity":
                r_rev = row_map["Total Revenue"]
                r_gbi = row_map["GBI"]
                r_rec = row_map["REC"]
                r_exp = row_map["Total Expenditure"]
                cell.value = f"={col_letter}{r_rev}+{col_letter}{r_gbi}+{col_letter}{r_rec}-{col_letter}{r_exp}"
            elif label == "Accumulated Profit/Loss":
                r_pft = row_map["Profit for Tax"]
                if idx == 0:
                    cell.value = f"={col_letter}{r_pft}"
                else:
                    prev_col = get_column_letter(col_idx - 1)
                    cell.value = f"={prev_col}{current_row}+{col_letter}{r_pft}"
            elif label == "Profit for Tax":
                r_pde = row_map["Profit due to sale of electricity"]
                cell.value = f"=IF({col_letter}2=1, 55.02, {col_letter}{r_pde})"
            elif label == "Corporate Tax":
                r_80 = row_map["80 IA"]
                r_pt = row_map["Profit for Tax"]
                cell.value = f"=IF({col_letter}{r_80}=1, 0, ROUND({col_letter}{r_pt}*B{current_row}, 2))"
            elif label == "MAT Credit Year Wise":
                r_ct = row_map["Corporate Tax"]
                r_mat = row_map["MAT"]
                cell.value = f"=IF({col_letter}{r_ct}>{col_letter}{r_mat}, 0, {col_letter}{r_mat}-{col_letter}{r_ct})"
            elif label == "Tax Payable Without MAT Credit":
                r_ct = row_map["Corporate Tax"]
                r_mat = row_map["MAT"]
                cell.value = f"=MAX({col_letter}{r_ct},{col_letter}{r_mat})"
            elif label == "Tax Payable After MAT Credit":
                r_tpw = row_map["Tax Payable Without MAT Credit"]
                r_mcu = row_map["MAT Credit Usage"]
                cell.value = f"={col_letter}{r_tpw}-{col_letter}{r_mcu}"
                # Apply accounting double underlines to last row!
                cell.border = border_double_bottom
            else:
                # Regular inputs / arrays
                if isinstance(source, list):
                    cell.value = source[idx]
                    
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    # Overwrite the border of the last row label/detail to double bottom border
    ws.cell(row=current_row-1, column=1).border = border_double_bottom
    ws.cell(row=current_row-1, column=2).border = border_double_bottom

def main():
    path_main = "c:/Users/Asus 15 Aspire/Desktop/another/Financial_Analysis_Model.xlsx"
    path_standalone = "c:/Users/Asus 15 Aspire/Desktop/another/Tax_Calculation.xlsx"

    # 1. Create standalone Tax_Calculation.xlsx first
    try:
        wb_sa = openpyxl.Workbook()
        ws_sa = wb_sa.active
        ws_sa.title = "Tax Calculation"
        build_tax_sheet(ws_sa)
        wb_sa.save(path_standalone)
        print(f"SUCCESS: Created standalone Tax Calculation sheet at: {path_standalone}")
    except Exception as e:
        print(f"ERROR: Failed to save standalone workbook: {e}")

    # 2. Update existing Financial_Analysis_Model.xlsx
    if os.path.exists(path_main):
        try:
            wb = openpyxl.load_workbook(path_main)
            if "Tax Calculation" in wb.sheetnames:
                del wb["Tax Calculation"]
            ws_new = wb.create_sheet("Tax Calculation", 1)  # Insert as sheet 2
            build_tax_sheet(ws_new)
            wb.save(path_main)
            print(f"SUCCESS: Added 'Tax Calculation' sheet to existing model at: {path_main}")
        except PermissionError:
            print(f"WARNING: Could not update '{path_main}' because it is open in Excel. Please close it and rerun the script.")
        except Exception as e:
            print(f"ERROR: Failed to update main workbook: {e}")
    else:
        print(f"Main workbook not found at: {path_main}")

if __name__ == "__main__":
    main()

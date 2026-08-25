import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

def create_assumptions_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumptions & Inputs"

    # Set gridlines visible
    ws.views.sheetView[0].showGridLines = True

    # Styling helpers
    font_title = Font(name="Calibri", size=14, bold=True)
    font_section = Font(name="Calibri", size=11, bold=True)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # Border setup
    thin_border_side = Side(border_style="thin", color="7F7F7F")
    border_thin = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Fill colors
    fill_title = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")     # Medium gray
    fill_section = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")   # Light gray for section header
    fill_input = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")     # Highlighted gray for input cells (as in image)

    # Margins and dimensions
    ws.column_dimensions["A"].width = 3   # Small margin column
    ws.column_dimensions["B"].width = 50  # Description
    ws.column_dimensions["C"].width = 15  # Unit / Detail
    ws.column_dimensions["D"].width = 20  # Value

    # Helper function to style a cell range (useful for merged cells)
    def style_range(ws, cell_range, font=None, alignment=None, fill=None, border=None):
        for row in ws[cell_range]:
            for cell in row:
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill
                if border:
                    cell.border = border

    current_row = 2

    # 1. Main Title
    ws.merge_cells("B2:D2")
    ws["B2"] = "Assumptions & Inputs"
    style_range(ws, "B2:D2", font=font_title, alignment=align_center, fill=fill_title, border=border_thin)
    ws.row_dimensions[2].height = 30
    current_row += 1

    # Empty spacer row
    ws.row_dimensions[current_row].height = 10
    current_row += 1

    # Define sections and their content
    # Format of each row: (Description, Unit/Detail, Value, is_input, number_format)
    # If Value is a string starting with "=", it is treated as a formula.
    sections_data = [
        # SECTION 1: Basic Assumptions
        {
            "title": "Basic Assumptions",
            "rows": [
                ("No.of Hours per day", "", 24, False, "#,##0"),
                ("No. of days in a year", "days", 365, False, "#,##0"),
                ("1 MW =", "KW", 1000, False, "#,##0"),
                ("No.of units per one crore", "", 10000000, False, "[>=10000000]##\\,##\\,##\\,##0;[>=100000]##\\,##\\,##0;##,##0")
            ]
        },
        # SECTION 2: Assumptions of Project
        {
            "title": "Assumptions of Project",
            "rows": [
                ("No.of days in a year", "days", 0.11, True, "0.00"),
                ("", "", "", True, ""), # Empty input row as per the image
                ("Life of Plant", "Years", 25, True, "#,##0"),
                ("Installed Capacity", "MW", 100.00, False, "0.00"),
                ("Number of units", "No.", 1, False, "#,##0"),
                ("Total Capacity", "", "=D17*D18", False, "#,##0"), # Hardcoded references: D17 (Installed Capacity) * D18 (Number of units)
                ("Gross Saleable Capacity", "MW", 100.00, False, "0.00"),
                ("Selling Price", "Rs", 6.30, False, "0.00"),
                ("Wheeling Charges per unit", "Rs", 0.52, False, "0.00"),
                ("Transmission loss ( 33 kv ) ( Appx)+ Wheeling loss", "Rs", 0.24, False, "0.00"),
                ("Transmission Charges per unit", "Rs", 0.50, False, "0.00"),
                ("Elec Duty TANGENDCO + SLDC+Operating charges", "Rs", 0.33, False, "0.00"),
                ("Base Tariff for UMPPL", "Per Unit", "=D21-D22-D23-D24-D25", False, "0.00"), # D21 (Selling Price) - D22 - D23 - D24 - D25
                ("PLF", "%", 0.2850, True, "0.00%"),
                ("System Losses (%)", "%", 0.0100, True, "0.00%")
            ]
        },
        # SECTION 3: Working Capital Norms & Escalations
        {
            "title": "Working Capital Norms & Escalations",
            "rows": [
                ("O&M and Insurance Expenses", "Month", 2, True, "#,##0"),
                ("O & M Expenses and Insurance (Rs Cr / MW)", "", 0.140, True, "0.000"),
                ("Assumed annual O&M cost increment", "", 0.0500, True, "0.00%"),
                ("Free O&M", "Years", 2.00, True, "0.00")
            ]
        },
        # SECTION 4: Debt
        {
            "title": "Debt",
            "rows": [
                ("Main debt", "Quarters", 60, True, "#,##0"),
                ("Moratorium", "Month", 6, True, "#,##0"),
                ("Installments per year (quarterly)", "", 4, True, "#,##0"),
                ("Repayment Period", "Years", "=D40/D42", True, "#,##0"), # Main debt quarters / Installments per year
                ("Interest on Main debt", "", 0.0850, True, "0.00%")
            ]
        },
        # SECTION 5: Salvage Value
        {
            "title": "Salvage Value",
            "rows": [
                ("Salvage/ Residual Value (Rs Crores)", 0.10, 78.87, False, "0.00"), # 10% in middle column, 78.87 in right
                ("Depreciation Rate For Years", 10.00, 0.07, False, "0%"), # 10.00 in middle column, 7% in right
                ("Thereafter Equal spread of balance value", "", "", False, "")
            ]
        },
        # SECTION 6: Income Tax
        {
            "title": "Income Tax",
            "rows": [
                ("Corporate Tax", "", 0.2600, True, "0.00%"),
                ("MAT", "", 0.2096, True, "0.00%")
            ]
        },
        # SECTION 7: Insurance and Misc
        {
            "title": "Insurance and Misc",
            "rows": [
                ("% of WDV", "", 0.0050, True, "0.00%"),
                ("Misc", "", 0.0050, True, "0.0%")
            ]
        }
    ]

    for section in sections_data:
        # Write section header
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=2, value=section["title"])
        style_range(ws, f"B{current_row}:D{current_row}", font=font_section, alignment=align_center, fill=fill_section, border=border_thin)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Write section rows
        for row_data in section["rows"]:
            desc, unit, val, is_input, num_format = row_data
            
            # Col B: Description
            c_desc = ws.cell(row=current_row, column=2, value=desc)
            c_desc.font = font_regular
            c_desc.alignment = align_left
            c_desc.border = border_thin
            
            # Col C: Unit / Middle Detail
            c_unit = ws.cell(row=current_row, column=3, value=unit)
            c_unit.font = font_regular
            c_unit.alignment = align_center
            c_unit.border = border_thin
            
            # If unit is a number, apply formatting
            if isinstance(unit, (int, float)):
                if unit == 0.10:
                    c_unit.number_format = "0%"
                elif unit == 10.00:
                    c_unit.number_format = "0.00"
                    c_unit.alignment = align_right

            # Col D: Value
            c_val = ws.cell(row=current_row, column=4, value=val)
            c_val.font = font_regular
            c_val.alignment = align_right
            c_val.border = border_thin
            
            if num_format:
                c_val.number_format = num_format
            
            # Apply input cell coloring if it's an input row
            if is_input:
                c_val.fill = fill_input
            
            # Special case for empty description but grey value cell (row 12 in excel, but dynamic here)
            if desc == "" and unit == "" and val == "":
                c_val.fill = fill_input
                
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Spacer row after section
        ws.row_dimensions[current_row].height = 10
        current_row += 1

    # Adjust formulas row index to be dynamically accurate!
    # Let's verify row numbers of formula cells:
    # Installed capacity is row 14, Number of units is row 15. Total capacity (row 16) should be =D14*D15
    # Let's inspect where they actually landed and overwrite formulas with exact row coordinates.
    row_installed_capacity = None
    row_num_units = None
    row_total_capacity = None
    
    row_selling_price = None
    row_wheeling_charges = None
    row_trans_loss = None
    row_trans_charges = None
    row_elec_duty = None
    row_base_tariff = None
    
    row_main_debt = None
    row_installments = None
    row_repayment_period = None

    for r in range(1, current_row):
        val_b = ws.cell(row=r, column=2).value
        if val_b == "Installed Capacity":
            row_installed_capacity = r
        elif val_b == "Number of units":
            row_num_units = r
        elif val_b == "Total Capacity":
            row_total_capacity = r
        elif val_b == "Selling Price":
            row_selling_price = r
        elif val_b == "Wheeling Charges per unit":
            row_wheeling_charges = r
        elif val_b == "Transmission loss ( 33 kv ) ( Appx)+ Wheeling loss":
            row_trans_loss = r
        elif val_b == "Transmission Charges per unit":
            row_trans_charges = r
        elif val_b == "Elec Duty TANGENDCO + SLDC+Operating charges":
            row_elec_duty = r
        elif val_b == "Base Tariff for UMPPL":
            row_base_tariff = r
        elif val_b == "Main debt":
            row_main_debt = r
        elif val_b == "Installments per year (quarterly)":
            row_installments = r
        elif val_b == "Repayment Period":
            row_repayment_period = r

    # Write corrected formulas using exact row coordinates!
    if row_total_capacity:
        ws.cell(row=row_total_capacity, column=4, value=f"=D{row_installed_capacity}*D{row_num_units}")
    if row_base_tariff:
        ws.cell(row=row_base_tariff, column=4, value=f"=D{row_selling_price}-D{row_wheeling_charges}-D{row_trans_loss}-D{row_trans_charges}-D{row_elec_duty}")
    if row_repayment_period:
        ws.cell(row=row_repayment_period, column=4, value=f"=D{row_main_debt}/D{row_installments}")

    # Output file path
    output_path = "c:/Users/Asus 15 Aspire/Desktop/another/Assumptions_and_Inputs.xlsx"
    wb.save(output_path)
    print(f"Assumptions & Inputs Excel file generated successfully at: {output_path}")

if __name__ == "__main__":
    create_assumptions_excel()

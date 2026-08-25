import openpyxl

file_path = r"C:\Users\Asus 15 Aspire\Downloads\Error_in_Formula_Original.xlsx"
wb_v = openpyxl.load_workbook(file_path, data_only=True)
ws = wb_v.active

# Diesel columns: G(7), L(12), Q(17), V(22), AA(27), AF(32), AK(37), AP(42), AU(47), AZ(52), BE(57), BJ(62), BO(67), BT(72), BY(77), CD(82), CI(87), CN(92), CS(97), CX(102), DC(107), DH(112), DM(117), DR(122), DW(127), EB(132), EG(137), EL(142), EQ(147), EV(152), FA(157)
diesel_cols = [7 + i*5 for i in range(31)]

non_numeric_cells = []

# Inspect all cells in data rows 8 to 46
for r in range(8, 47):
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=r, column=c).value
        
        # Check if value is string containing '-', ';', space, or non-numeric
        if val is not None and isinstance(val, str):
            val_str = str(val).strip()
            col_letter = openpyxl.utils.get_column_letter(c)
            header_col = ws.cell(row=6, column=c).value or ws.cell(row=5, column=c).value or ""
            vehicle_no = ws.cell(row=5, column=c-c%5 if c%5!=0 else c-5).value if c > 6 else ""
            
            non_numeric_cells.append({
                'row': r,
                'col': c,
                'col_letter': col_letter,
                'cell_ref': f"{col_letter}{r}",
                'value': repr(val),
                'header': header_col,
                'is_diesel_col': c in diesel_cols
            })

print(f"Total non-numeric string cells found in data area (rows 8-46): {len(non_numeric_cells)}")
print("\n--- DIESEL COLUMNS NON-NUMERIC/DASH/SEMICOLON CELLS ---")
diesel_non_num = [item for item in non_numeric_cells if item['is_diesel_col']]
print(f"Total non-numeric entries in DIESEL columns: {len(diesel_non_num)}")

for item in diesel_non_num:
    print(f"Cell {item['cell_ref']} (Row {item['row']}, Col {item['col_letter']}): Value={item['value']} | Header='{item['header']}'")

print("\n--- OTHER COLUMNS NON-NUMERIC/DASH/SEMICOLON CELLS (First 30) ---")
for item in non_numeric_cells[:30]:
    print(f"Cell {item['cell_ref']} (Row {item['row']}, Col {item['col_letter']}): Value={item['value']}")
